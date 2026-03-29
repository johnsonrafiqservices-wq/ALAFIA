from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from datetime import datetime, timedelta
from decimal import Decimal
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

from .models import Budget, BudgetItem, Expense, ExpenseCategory
from .forms import BudgetForm, BudgetItemForm, ExpenseForm, ExpenseCategoryForm, ExpenseApprovalForm


@login_required
def budget_dashboard(request):
    """Main budget and expense dashboard"""
    # Get active budget
    active_budget = Budget.objects.filter(status='active').first()
    
    # Get current month stats
    now = timezone.now()
    current_month_start = datetime(now.year, now.month, 1).date()
    
    # Expense stats
    total_expenses = Expense.objects.filter(status='approved').aggregate(
        total=Sum('amount'))['total'] or Decimal('0')
    
    month_expenses = Expense.objects.filter(
        status='approved',
        expense_date__gte=current_month_start
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    pending_expenses = Expense.objects.filter(status='pending').count()
    
    # Recent expenses
    recent_expenses = Expense.objects.all()[:10]
    
    # Category breakdown
    category_stats = ExpenseCategory.objects.filter(
        is_active=True
    ).annotate(
        total_spent=Sum('expenses__amount', filter=Q(expenses__status='approved')),
        expense_count=Count('expenses', filter=Q(expenses__status='approved'))
    ).order_by('-total_spent')[:5]
    
    # Budget summary
    all_budgets = Budget.objects.all()[:5]
    
    # Get categories and budget items for modal forms
    categories = ExpenseCategory.objects.filter(is_active=True)
    budget_items = BudgetItem.objects.select_related('budget', 'category').filter(
        budget__status='active'
    )
    
    context = {
        'active_budget': active_budget,
        'total_expenses': total_expenses,
        'month_expenses': month_expenses,
        'pending_expenses': pending_expenses,
        'recent_expenses': recent_expenses,
        'category_stats': category_stats,
        'all_budgets': all_budgets,
        'categories': categories,
        'budget_items': budget_items,
    }
    
    return render(request, 'budget/dashboard.html', context)


@login_required
def budget_list(request):
    """List all budgets with filters"""
    budgets = Budget.objects.all()
    
    # Filters
    status_filter = request.GET.get('status')
    period_type_filter = request.GET.get('period_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    search_query = request.GET.get('search')
    
    if status_filter:
        budgets = budgets.filter(status=status_filter)
    
    if period_type_filter:
        budgets = budgets.filter(period_type=period_type_filter)
    
    if date_from:
        budgets = budgets.filter(start_date__gte=date_from)
    
    if date_to:
        budgets = budgets.filter(end_date__lte=date_to)
    
    if search_query:
        budgets = budgets.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Calculate totals
    total_budget_amount = budgets.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    # Calculate total spent and remaining across all filtered budgets
    total_spent = Decimal('0')
    total_remaining = Decimal('0')
    for budget in budgets:
        total_spent += budget.get_spent_amount()
        total_remaining += budget.get_remaining_amount()
    
    # Get choices for filters
    status_choices = Budget.STATUS_CHOICES
    period_choices = Budget.PERIOD_CHOICES
    
    context = {
        'budgets': budgets,
        'status_choices': status_choices,
        'period_choices': period_choices,
        'status_filter': status_filter,
        'period_type_filter': period_type_filter,
        'date_from': date_from,
        'date_to': date_to,
        'search_query': search_query,
        'total_budget_amount': total_budget_amount,
        'total_spent': total_spent,
        'total_remaining': total_remaining,
    }
    
    return render(request, 'budget/budget_list.html', context)


@login_required
def budget_create(request):
    """Create new budget"""
    if request.method == 'POST':
        form = BudgetForm(request.POST)
        if form.is_valid():
            budget = form.save(commit=False)
            budget.created_by = request.user
            budget.save()
            messages.success(request, f'Budget "{budget.name}" created successfully!')
            return redirect('budget:budget_detail', pk=budget.pk)
    else:
        form = BudgetForm()
    
    context = {
        'form': form,
        'title': 'Create New Budget',
    }
    
    return render(request, 'budget/budget_form.html', context)


@login_required
def budget_detail(request, pk):
    """View budget details"""
    budget = get_object_or_404(Budget, pk=pk)
    
    # Calculate stats
    allocated_amount = budget.get_allocated_amount()
    spent_amount = budget.get_spent_amount()
    remaining_amount = budget.get_remaining_amount()
    utilization_pct = budget.get_utilization_percentage()
    
    # Get budget items with stats
    budget_items = budget.items.all()
    
    # Get categories for modal form
    categories = ExpenseCategory.objects.filter(is_active=True)
    
    context = {
        'budget': budget,
        'allocated_amount': allocated_amount,
        'spent_amount': spent_amount,
        'remaining_amount': remaining_amount,
        'utilization_pct': utilization_pct,
        'budget_items': budget_items,
        'categories': categories,
    }
    
    return render(request, 'budget/budget_detail.html', context)


@login_required
def budget_edit(request, pk):
    """Edit existing budget"""
    budget = get_object_or_404(Budget, pk=pk)
    
    if request.method == 'POST':
        form = BudgetForm(request.POST, instance=budget)
        if form.is_valid():
            form.save()
            messages.success(request, f'Budget "{budget.name}" updated successfully!')
            return redirect('budget:budget_detail', pk=budget.pk)
    else:
        form = BudgetForm(instance=budget)
    
    context = {
        'form': form,
        'budget': budget,
        'title': f'Edit Budget: {budget.name}',
    }
    
    return render(request, 'budget/budget_form.html', context)


@login_required
def budget_item_create(request, budget_pk):
    """Add budget item to budget"""
    budget = get_object_or_404(Budget, pk=budget_pk)
    
    if request.method == 'POST':
        form = BudgetItemForm(request.POST)
        if form.is_valid():
            budget_item = form.save(commit=False)
            budget_item.budget = budget
            budget_item.save()
            messages.success(request, 'Budget item added successfully!')
            return redirect('budget:budget_detail', pk=budget.pk)
    else:
        form = BudgetItemForm()
    
    context = {
        'form': form,
        'budget': budget,
        'title': f'Add Item to {budget.name}',
    }
    
    return render(request, 'budget/budget_item_form.html', context)


@login_required
def expense_list(request):
    """List all expenses with filters"""
    expenses = Expense.objects.all()
    
    # Filters
    status_filter = request.GET.get('status')
    category_filter = request.GET.get('category')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    payment_method_filter = request.GET.get('payment_method')
    search_query = request.GET.get('search')
    
    if status_filter:
        expenses = expenses.filter(status=status_filter)
    
    if category_filter:
        expenses = expenses.filter(category_id=category_filter)
    
    if date_from:
        expenses = expenses.filter(expense_date__gte=date_from)
    
    if date_to:
        expenses = expenses.filter(expense_date__lte=date_to)
    
    if payment_method_filter:
        expenses = expenses.filter(payment_method=payment_method_filter)
    
    if search_query:
        expenses = expenses.filter(
            Q(description__icontains=search_query) |
            Q(vendor_name__icontains=search_query) |
            Q(reference_number__icontains=search_query)
        )
    
    # Calculate total amount for filtered expenses
    total_amount = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Get categories for filter
    categories = ExpenseCategory.objects.filter(is_active=True)
    
    # Get payment method choices for filter
    payment_methods = Expense.PAYMENT_METHOD_CHOICES
    
    # Get status choices for filter
    status_choices = Expense.STATUS_CHOICES
    
    context = {
        'expenses': expenses,
        'categories': categories,
        'payment_methods': payment_methods,
        'status_choices': status_choices,
        'status_filter': status_filter,
        'category_filter': category_filter,
        'date_from': date_from,
        'date_to': date_to,
        'payment_method_filter': payment_method_filter,
        'search_query': search_query,
        'total_amount': total_amount,
    }
    
    return render(request, 'budget/expense_list.html', context)


@login_required
def expense_create(request):
    """Create new expense"""
    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.submitted_by = request.user
            expense.save()
            messages.success(request, 'Expense submitted successfully and is pending approval!')
            return redirect('budget:expense_detail', pk=expense.pk)
    else:
        form = ExpenseForm()
    
    context = {
        'form': form,
        'title': 'Add New Expense',
    }
    
    return render(request, 'budget/expense_form.html', context)


@login_required
def expense_detail(request, pk):
    """View expense details"""
    expense = get_object_or_404(Expense, pk=pk)
    
    context = {
        'expense': expense,
    }
    
    return render(request, 'budget/expense_detail.html', context)


@login_required
def expense_edit(request, pk):
    """Edit expense (only if pending)"""
    expense = get_object_or_404(Expense, pk=pk)
    
    # Only allow editing of pending expenses
    if expense.status != 'pending':
        messages.error(request, 'Only pending expenses can be edited.')
        return redirect('budget:expense_detail', pk=expense.pk)
    
    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, 'Expense updated successfully!')
            return redirect('budget:expense_detail', pk=expense.pk)
    else:
        form = ExpenseForm(instance=expense)
    
    context = {
        'form': form,
        'expense': expense,
        'title': 'Edit Expense',
    }
    
    return render(request, 'budget/expense_form.html', context)


@login_required
def expense_approve(request, pk):
    """Approve or reject expense"""
    expense = get_object_or_404(Expense, pk=pk)
    
    # Only allow approval of pending expenses
    if expense.status != 'pending':
        messages.error(request, 'Only pending expenses can be approved or rejected.')
        return redirect('budget:expense_detail', pk=expense.pk)
    
    if request.method == 'POST':
        status = request.POST.get('status')
        approval_notes = request.POST.get('approval_notes', '')
        rejection_reason = request.POST.get('rejection_reason', '')
        
        if not status:
            messages.error(request, 'Please select Approve or Reject.')
            return redirect('budget:expense_approve', pk=expense.pk)
        
        if status == 'approved':
            expense.status = 'approved'
            expense.approved_by = request.user
            expense.approved_date = timezone.now()
            if approval_notes:
                expense.notes = (expense.notes + '\n\n' + approval_notes) if expense.notes else approval_notes
            expense.save()
            messages.success(request, f'Expense "{expense.description}" approved successfully!')
        elif status == 'rejected':
            if not rejection_reason:
                messages.error(request, 'Rejection reason is required.')
                return redirect('budget:expense_approve', pk=expense.pk)
            
            expense.status = 'rejected'
            expense.approved_by = request.user
            expense.approved_date = timezone.now()
            expense.rejection_reason = rejection_reason
            if approval_notes:
                expense.notes = (expense.notes + '\n\n' + approval_notes) if expense.notes else approval_notes
            expense.save()
            messages.success(request, f'Expense "{expense.description}" rejected.')
        
        return redirect('budget:expense_detail', pk=expense.pk)
    
    context = {
        'expense': expense,
    }
    
    return render(request, 'budget/expense_approve.html', context)


@login_required
def category_list(request):
    """List all expense categories"""
    categories = ExpenseCategory.objects.all()
    
    context = {
        'categories': categories,
    }
    
    return render(request, 'budget/category_list.html', context)


@login_required
def category_create(request):
    """Create new expense category"""
    if request.method == 'POST':
        form = ExpenseCategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, f'Category "{category.name}" created successfully!')
            return redirect('budget:category_list')
    else:
        form = ExpenseCategoryForm()
    
    context = {
        'form': form,
        'title': 'Create Expense Category',
    }
    
    return render(request, 'budget/category_form.html', context)


# AJAX Views
@login_required
def expense_create_ajax(request):
    """AJAX-only expense creation view"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    # Check if request is AJAX
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    form = ExpenseForm(request.POST, request.FILES)
    if form.is_valid():
        expense = form.save(commit=False)
        expense.submitted_by = request.user
        expense.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Expense "{expense.description}" submitted successfully and is pending approval!',
            'expense_id': expense.pk,
            'redirect_url': reverse('budget:expense_detail', kwargs={'pk': expense.pk})
        })
    else:
        # Return form errors for client-side display
        errors = {}
        for field, error_list in form.errors.items():
            errors[field] = error_list
        
        return JsonResponse({
            'success': False,
            'errors': errors,
            'message': 'Please correct the errors below and try again.'
        }, status=400)


@login_required
def budget_create_ajax(request):
    """AJAX-only budget creation view"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    # Check if request is AJAX
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    form = BudgetForm(request.POST)
    if form.is_valid():
        budget = form.save(commit=False)
        budget.created_by = request.user
        budget.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Budget "{budget.name}" created successfully!',
            'budget_id': budget.pk,
            'redirect_url': reverse('budget:budget_detail', kwargs={'pk': budget.pk})
        })
    else:
        # Return form errors for client-side display
        errors = {}
        for field, error_list in form.errors.items():
            errors[field] = error_list
        
        return JsonResponse({
            'success': False,
            'errors': errors,
            'message': 'Please correct the errors below and try again.'
        }, status=400)


@login_required
def budget_item_create_ajax(request, budget_pk):
    """AJAX-only budget item creation view (single item - deprecated, use add_items)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    # Check if request is AJAX
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    budget = get_object_or_404(Budget, pk=budget_pk)
    
    form = BudgetItemForm(request.POST)
    if form.is_valid():
        budget_item = form.save(commit=False)
        budget_item.budget = budget
        budget_item.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Budget item for "{budget_item.category.name}" added successfully!',
            'budget_item_id': budget_item.pk,
            'redirect_url': reverse('budget:budget_detail', kwargs={'pk': budget.pk})
        })
    else:
        # Return form errors for client-side display
        errors = {}
        for field, error_list in form.errors.items():
            errors[field] = error_list
        
        return JsonResponse({
            'success': False,
            'errors': errors,
            'message': 'Please correct the errors below and try again.'
        }, status=400)


@login_required
def budget_items_create_multiple_ajax(request, budget_pk):
    """AJAX-only multiple budget items creation"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    budget = get_object_or_404(Budget, pk=budget_pk)
    
    try:
        data = json.loads(request.body)
        items = data.get('items', [])
        
        if not items:
            return JsonResponse({
                'success': False,
                'message': 'No items provided'
            }, status=400)
        
        created_items = []
        errors = []
        
        for idx, item in enumerate(items):
            try:
                category_id = item.get('category')
                if not category_id:
                    errors.append(f'Row {idx + 1}: Category is required')
                    continue
                
                category = ExpenseCategory.objects.get(pk=category_id)
                
                amount = item.get('allocated_amount', 0)
                if not amount or amount == 0:
                    errors.append(f'Row {idx + 1}: Amount must be greater than 0')
                    continue
                
                allocated_amount = Decimal(str(amount))
                notes = item.get('description', '')  # Frontend sends 'description', we save as 'notes'
                
                budget_item = BudgetItem.objects.create(
                    budget=budget,
                    category=category,
                    allocated_amount=allocated_amount,
                    notes=notes
                )
                created_items.append(budget_item)
            except ExpenseCategory.DoesNotExist:
                errors.append(f'Row {idx + 1}: Invalid category ID {category_id}')
            except (ValueError, TypeError) as e:
                errors.append(f'Row {idx + 1}: Invalid amount format - {str(e)}')
            except Exception as e:
                errors.append(f'Row {idx + 1}: {str(e)}')
        
        if errors:
            return JsonResponse({
                'success': False,
                'message': f'Created {len(created_items)} items with {len(errors)} errors',
                'errors': errors
            }, status=400)
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully added {len(created_items)} budget items!',
            'created_count': len(created_items),
            'redirect_url': reverse('budget:budget_detail', kwargs={'pk': budget.pk})
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data'
        }, status=400)


@login_required
def budget_items_import_excel(request, budget_pk):
    """Import budget items from Excel file"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    budget = get_object_or_404(Budget, pk=budget_pk)
    
    if 'excel_file' not in request.FILES:
        return JsonResponse({
            'success': False,
            'message': 'No file uploaded'
        }, status=400)
    
    excel_file = request.FILES['excel_file']
    
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        
        created_items = []
        errors = []
        
        # Skip header row
        for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:  # Skip empty rows
                continue
            
            try:
                # Validate category
                if not row[0]:
                    errors.append(f'Row {idx}: Category is required')
                    continue
                    
                category_name = str(row[0]).strip()
                
                # Validate amount
                if len(row) < 2 or not row[1]:
                    errors.append(f'Row {idx}: Amount is required')
                    continue
                
                try:
                    amount = Decimal(str(row[1]))
                    if amount <= 0:
                        errors.append(f'Row {idx}: Amount must be greater than 0')
                        continue
                except (ValueError, TypeError):
                    errors.append(f'Row {idx}: Invalid amount "{row[1]}" - must be a number')
                    continue
                
                notes = str(row[2]) if len(row) > 2 and row[2] else ''
                
                # Find category by name
                category = ExpenseCategory.objects.filter(name__iexact=category_name).first()
                
                if not category:
                    available_categories = ', '.join(ExpenseCategory.objects.filter(is_active=True).values_list('name', flat=True)[:5])
                    errors.append(f'Row {idx}: Category "{category_name}" not found. Available: {available_categories}...')
                    continue
                
                budget_item = BudgetItem.objects.create(
                    budget=budget,
                    category=category,
                    allocated_amount=amount,
                    notes=notes
                )
                created_items.append(budget_item)
                
            except Exception as e:
                errors.append(f'Row {idx}: Unexpected error - {str(e)}')
        
        if not created_items and errors:
            return JsonResponse({
                'success': False,
                'message': 'No items could be imported',
                'errors': errors
            }, status=400)
        
        message = f'Successfully imported {len(created_items)} budget items'
        if errors:
            message += f' (with {len(errors)} errors)'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'created_count': len(created_items),
            'errors': errors if errors else None,
            'redirect_url': reverse('budget:budget_detail', kwargs={'pk': budget.pk})
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error processing Excel file: {str(e)}'
        }, status=500)


@login_required
def budget_items_download_template(request, budget_pk):
    """Download Excel template for budget items import"""
    budget = get_object_or_404(Budget, pk=budget_pk)
    
    # Create workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Budget Items'
    
    # Define headers
    headers = ['Category', 'Amount', 'Description']
    
    # Style headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Add sample data with categories
    categories = ExpenseCategory.objects.filter(is_active=True)[:10]
    for idx, category in enumerate(categories, start=2):
        sheet.cell(row=idx, column=1, value=category.name)
        sheet.cell(row=idx, column=2, value=0)
        sheet.cell(row=idx, column=3, value=f'Notes for {category.name}')
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 40
    
    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=budget_items_template_{budget.name.replace(" ", "_")}.xlsx'
    
    return response


@login_required
def categories_list_ajax(request):
    """Get list of active categories for AJAX requests"""
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    categories = ExpenseCategory.objects.filter(is_active=True).values('id', 'name', 'icon', 'color')
    
    return JsonResponse({
        'categories': list(categories)
    })
