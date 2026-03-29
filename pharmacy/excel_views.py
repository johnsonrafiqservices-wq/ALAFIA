"""
Excel Import Views for Pharmacy
Handles bulk import of medications and batches from Excel files
"""
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from django.utils import timezone
from datetime import datetime
from decimal import Decimal
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Medication, Batch


@login_required
@require_POST
def import_medications(request):
    """Import medications from Excel file"""
    try:
        excel_file = request.FILES.get('excel_file')
        skip_duplicates = request.POST.get('skip_duplicates') == 'on'
        
        if not excel_file:
            return JsonResponse({'success': False, 'message': 'No file uploaded'}, status=400)
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error reading file: {str(e)}'}, status=400)
        
        # Process import
        imported = 0
        skipped = 0
        errors = 0
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        required = ['name', 'generic_name', 'category', 'dosage_form', 'strength', 'unit_of_measure']
        
        # Validate headers
        missing = [col for col in required if col not in headers]
        if missing:
            return JsonResponse({
                'success': False,
                'message': f'Missing columns: {", ".join(missing)}'
            })
        
        # Get column indices
        col_idx = {header: idx for idx, header in enumerate(headers)}
        
        # Process rows
        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    name = row[col_idx['name']]
                    if not name:
                        continue
                    
                    # Check if exists
                    if Medication.objects.filter(name=name).exists():
                        if skip_duplicates:
                            skipped += 1
                            continue
                    
                    # Create medication
                    Medication.objects.create(
                        name=name,
                        generic_name=row[col_idx['generic_name']],
                        category=row[col_idx['category']],
                        dosage_form=row[col_idx['dosage_form']],
                        strength=row[col_idx['strength']],
                        unit_of_measure=row[col_idx['unit_of_measure']],
                        is_active=True
                    )
                    imported += 1
                except Exception as e:
                    errors += 1
                    print(f"Error importing row: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': 'Import completed',
            'total': imported + skipped + errors,
            'imported': imported,
            'skipped': skipped,
            'errors': errors
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Import error: {str(e)}'
        }, status=500)


@login_required
@require_POST
def import_batches(request):
    """Import batches from Excel file"""
    try:
        excel_file = request.FILES.get('excel_file')
        skip_duplicates = request.POST.get('skip_duplicates') == 'on'
        
        if not excel_file:
            return JsonResponse({'success': False, 'message': 'No file uploaded'}, status=400)
        
        # Load workbook
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error reading file: {str(e)}'}, status=400)
        
        # Process import
        imported = 0
        skipped = 0
        errors = 0
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        required = ['medication_name', 'batch_number', 'quantity', 'manufacturing_date', 
                    'expiry_date', 'cost_price', 'selling_price']
        
        # Validate headers
        missing = [col for col in required if col not in headers]
        if missing:
            return JsonResponse({
                'success': False,
                'message': f'Missing columns: {", ".join(missing)}'
            })
        
        # Get column indices
        col_idx = {header: idx for idx, header in enumerate(headers)}
        
        # Process rows
        with transaction.atomic():
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    batch_number = row[col_idx['batch_number']]
                    if not batch_number:
                        continue
                    
                    # Find medication
                    medication_name = row[col_idx['medication_name']]
                    try:
                        medication = Medication.objects.get(name=medication_name, is_active=True)
                    except Medication.DoesNotExist:
                        errors += 1
                        continue
                    
                    # Check if batch exists
                    if Batch.objects.filter(batch_number=batch_number).exists():
                        if skip_duplicates:
                            skipped += 1
                            continue
                    
                    # Parse dates
                    mfg_date = row[col_idx['manufacturing_date']]
                    exp_date = row[col_idx['expiry_date']]
                    
                    if isinstance(mfg_date, str):
                        mfg_date = datetime.strptime(mfg_date, '%Y-%m-%d').date()
                    if isinstance(exp_date, str):
                        exp_date = datetime.strptime(exp_date, '%Y-%m-%d').date()
                    
                    # Create batch
                    quantity = int(row[col_idx['quantity']])
                    Batch.objects.create(
                        medication=medication,
                        batch_number=batch_number,
                        quantity=quantity,
                        quantity_remaining=quantity,
                        manufacturing_date=mfg_date,
                        expiry_date=exp_date,
                        cost_price=Decimal(str(row[col_idx['cost_price']])),
                        selling_price=Decimal(str(row[col_idx['selling_price']])),
                        is_active=True,
                        received_by=request.user
                    )
                    imported += 1
                except Exception as e:
                    errors += 1
                    print(f"Error importing batch: {str(e)}")
        
        return JsonResponse({
            'success': True,
            'message': 'Import completed',
            'total': imported + skipped + errors,
            'imported': imported,
            'skipped': skipped,
            'errors': errors
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Import error: {str(e)}'
        }, status=500)


@login_required
def download_medication_template(request):
    """Download Excel template for medication import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Medications"
    
    # Headers
    headers = ['name', 'generic_name', 'category', 'dosage_form', 'strength', 'unit_of_measure']
    
    # Style headers
    header_fill = PatternFill(start_color="1B5E96", end_color="1B5E96", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Sample data
    samples = [
        ['Paracetamol 500mg', 'Acetaminophen', 'Analgesics', 'Tablet', '500mg', 'tablets'],
        ['Amoxicillin 250mg', 'Amoxicillin', 'Antibiotics', 'Capsule', '250mg', 'capsules'],
        ['Ibuprofen 400mg', 'Ibuprofen', 'Anti-inflammatory', 'Tablet', '400mg', 'tablets'],
    ]
    
    for row_num, row_data in enumerate(samples, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column[0].column_letter].width = max_length + 2
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=medication_import_template.xlsx'
    wb.save(response)
    return response


@login_required
def download_batch_template(request):
    """Download Excel template for batch import"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Batches"
    
    # Headers
    headers = ['medication_name', 'batch_number', 'quantity', 'manufacturing_date', 
               'expiry_date', 'cost_price', 'selling_price']
    
    # Style headers
    header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Sample data
    samples = [
        ['Paracetamol 500mg', 'BATCH001', 1000, '2024-01-01', '2026-12-31', 50, 100],
        ['Amoxicillin 250mg', 'BATCH002', 500, '2024-02-01', '2026-01-31', 150, 250],
        ['Ibuprofen 400mg', 'BATCH003', 750, '2024-03-01', '2026-02-28', 80, 150],
    ]
    
    for row_num, row_data in enumerate(samples, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        ws.column_dimensions[column[0].column_letter].width = max_length + 2
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=batch_import_template.xlsx'
    wb.save(response)
    return response
