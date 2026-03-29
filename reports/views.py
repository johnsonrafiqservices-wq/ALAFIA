from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Avg, Q, Max, Min
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.core.cache import cache
from datetime import date, timedelta, datetime
from decimal import Decimal
import json
import time
from patients.models import Patient
from appointments.models import Appointment, Service
from billing.models import Invoice, Payment, InvoiceLineItem
from accounts.models import User
from .models import ReportAuditLog, ReportExport, ReportConfiguration
from .utils import ReportAuditMixin, create_report_export, get_report_performance_metrics

class ReportsView(ReportAuditMixin):
    """Base class for report views with audit logging"""
    pass

@login_required
def reports_dashboard(request):
    """Enhanced reports dashboard with audit logging"""
    start_time = time.time()
    audit_mixin = ReportAuditMixin()
    # Date range for reports (default to current month)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        start_date = date.today().replace(day=1)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = date.today()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Patient statistics
    total_patients = Patient.objects.filter(is_active=True).count()
    new_patients = Patient.objects.filter(
        registration_date__range=[start_date, end_date]
    ).count()
    
    # Appointment statistics
    total_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date]
    ).count()
    
    completed_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date],
        status='completed'
    ).count()
    
    # Revenue statistics
    total_revenue = Payment.objects.filter(
        payment_date__date__range=[start_date, end_date],
        status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Outstanding balance - calculate from unpaid/partially paid invoices
    unpaid_invoices = Invoice.objects.filter(
        status__in=['sent', 'overdue']
    )
    outstanding_balance = 0
    for invoice in unpaid_invoices:
        outstanding_balance += invoice.get_balance_due()
    
    # Service popularity
    popular_services = Service.objects.annotate(
        appointment_count=Count('appointment')
    ).order_by('-appointment_count')[:5]
    
    service_labels = []
    service_data = []
    for service in popular_services:
        service_labels.append(service.name)
        service_data.append(service.appointment_count)
    
    # Chart data for dashboard - 7 day revenue trend
    revenue_labels = []
    revenue_data = []
    for i in range(7):
        day = end_date - timedelta(days=i)
        revenue = Payment.objects.filter(
            payment_date__date=day,
            status='completed'
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        revenue_labels.insert(0, day.strftime('%m/%d'))
        revenue_data.insert(0, float(revenue))
    
    # Patient demographics
    gender_data = Patient.objects.values('gender').annotate(count=Count('id'))
    gender_labels = []
    gender_counts = []
    for item in gender_data:
        gender_labels.append(item['gender'])
        gender_counts.append(item['count'])
    
    # Department statistics
    from patients.models import Assessment
    physio_count = Assessment.objects.filter(
        department='physiotherapy',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    nutrition_count = Assessment.objects.filter(
        department='nutrition',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    general_count = Assessment.objects.filter(
        department='general',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    # Appointment status breakdown
    scheduled_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date],
        status='scheduled'
    ).count()
    
    cancelled_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date],
        status='cancelled'
    ).count()
    
    no_show_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date],
        status='no_show'
    ).count()
    
    # Calculate execution time and log activity
    execution_time = time.time() - start_time
    audit_mixin.log_report_activity(
        request=request,
        report_type='dashboard',
        report_name='Reports Dashboard',
        action='viewed',
        execution_time=execution_time,
        record_count=total_patients + total_appointments,
        parameters={
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat()
        }
    )
    
    # Get recent audit activities for dashboard
    recent_activities = ReportAuditLog.objects.filter(
        user=request.user
    ).order_by('-timestamp')[:10]
    
    # Get performance metrics
    performance_metrics = get_report_performance_metrics()
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_patients': total_patients,
        'new_patients': new_patients,
        'total_appointments': total_appointments,
        'completed_appointments': completed_appointments,
        'scheduled_appointments': scheduled_appointments,
        'cancelled_appointments': cancelled_appointments,
        'no_show_appointments': no_show_appointments,
        'total_revenue': total_revenue,
        'outstanding_balance': outstanding_balance,
        'popular_services': popular_services,
        'service_labels': json.dumps(service_labels),
        'service_data': json.dumps(service_data),
        'revenue_labels': json.dumps(revenue_labels),
        'revenue_data': json.dumps(revenue_data),
        'gender_labels': json.dumps(gender_labels),
        'gender_data': json.dumps(gender_counts),
        'physio_count': physio_count,
        'nutrition_count': nutrition_count,
        'general_count': general_count,
        'satisfaction_score': 0,  # Calculate from real data if needed
        'recent_activities': recent_activities,
        'performance_metrics': performance_metrics,
        'execution_time': round(execution_time, 2),
    }
    return render(request, 'reports/dashboard.html', context)

@login_required
def patient_reports(request):
    from patients.models import Assessment
    from calendar import monthrange as cal_monthrange

    # ── Filters ───────────────────────────────────────────────────
    date_range = request.GET.get('date_range', 'last_30_days')
    age_group  = request.GET.get('age_group', '')
    gender     = request.GET.get('gender', '')
    date_from  = request.GET.get('date_from', '')
    date_to    = request.GET.get('date_to', '')

    end_date = date.today()

    if date_range == 'custom' and date_from and date_to:
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date   = datetime.strptime(date_to,   '%Y-%m-%d').date()
        except ValueError:
            start_date = end_date - timedelta(days=30)
    elif date_range == 'last_3_months':
        start_date = end_date - timedelta(days=90)
    elif date_range == 'last_6_months':
        start_date = end_date - timedelta(days=180)
    elif date_range == 'last_year':
        start_date = end_date - timedelta(days=365)
    else:
        date_range = 'last_30_days'
        start_date = end_date - timedelta(days=30)

    # ── Base queryset ──────────────────────────────────────────────
    patients = Patient.objects.filter(is_active=True)
    if gender:
        patients = patients.filter(gender=gender)

    # ── Age group filter ───────────────────────────────────────────
    if age_group:
        age_filtered_ids = []
        for p in patients.only('id', 'date_of_birth'):
            try:
                age = p.get_age()
                if not isinstance(age, (int, float)):
                    continue
                age = int(age)
                if age_group == '0-18' and age <= 18:
                    age_filtered_ids.append(p.pk)
                elif age_group == '19-35' and 19 <= age <= 35:
                    age_filtered_ids.append(p.pk)
                elif age_group == '36-55' and 36 <= age <= 55:
                    age_filtered_ids.append(p.pk)
                elif age_group == '56+' and age >= 56:
                    age_filtered_ids.append(p.pk)
            except (ValueError, TypeError, AttributeError):
                continue
        patients = patients.filter(pk__in=age_filtered_ids)

    # ── Summary stats ──────────────────────────────────────────────
    total_patients = patients.count()
    new_patients   = patients.filter(registration_date__date__range=[start_date, end_date]).count()

    # Active = had at least one assessment in the selected period
    active_patients = patients.filter(
        assessments__assessment_date__date__range=[start_date, end_date]
    ).distinct().count()

    # Average age (skip "Unknown" strings)
    ages = []
    for p in patients.only('id', 'date_of_birth'):
        try:
            age = p.get_age()
            if isinstance(age, (int, float)):
                ages.append(int(age))
        except (ValueError, TypeError, AttributeError):
            continue
    average_age = round(sum(ages) / len(ages), 1) if ages else 0

    # ── Age distribution ───────────────────────────────────────────
    age_labels = ['0–18', '19–35', '36–55', '56+']
    age_data   = [0, 0, 0, 0]
    for p in patients.only('id', 'date_of_birth'):
        try:
            age = p.get_age()
            if not isinstance(age, (int, float)):
                continue
            age = int(age)
            if age <= 18:
                age_data[0] += 1
            elif age <= 35:
                age_data[1] += 1
            elif age <= 55:
                age_data[2] += 1
            else:
                age_data[3] += 1
        except (ValueError, TypeError, AttributeError):
            continue

    # ── Gender distribution (human-readable labels) ────────────────
    GENDER_MAP = {'M': 'Male', 'F': 'Female', 'O': 'Other', '': 'Not Specified'}
    gender_stats  = patients.values('gender').annotate(count=Count('id'))
    gender_labels = [GENDER_MAP.get(item['gender'] or '', item['gender'] or 'Not Specified') for item in gender_stats]
    gender_data   = [item['count'] for item in gender_stats]

    # ── Registration trend (last 6 months) ────────────────────────
    trend_labels = []
    trend_data   = []
    for i in range(5, -1, -1):
        m_ref = end_date.month - i
        y_ref = end_date.year
        while m_ref <= 0:
            m_ref += 12
            y_ref -= 1
        _, ml = cal_monthrange(y_ref, m_ref)
        ms = date(y_ref, m_ref, 1)
        me = date(y_ref, m_ref, ml)
        trend_labels.append(ms.strftime('%b %Y'))
        trend_data.append(Patient.objects.filter(registration_date__date__range=[ms, me]).count())

    # ── Insurance ──────────────────────────────────────────────────
    insurance_stats  = patients.values('insurance_provider').annotate(count=Count('id')).order_by('-count')[:5]
    insurance_labels = [i['insurance_provider'] for i in insurance_stats if i['insurance_provider']]
    insurance_data   = [i['count'] for i in insurance_stats if i['insurance_provider']]

    # ── Department stats (filtered by date range) ──────────────────
    physio_patients    = patients.filter(
        assessments__department='physiotherapy',
        assessments__assessment_date__date__range=[start_date, end_date]
    ).distinct().count()
    nutrition_patients = patients.filter(
        assessments__department='nutrition',
        assessments__assessment_date__date__range=[start_date, end_date]
    ).distinct().count()
    general_patients   = patients.filter(
        assessments__department='general',
        assessments__assessment_date__date__range=[start_date, end_date]
    ).distinct().count()

    dept_labels = ['Physiotherapy', 'Nutrition', 'General Medicine']
    dept_data   = [physio_patients, nutrition_patients, general_patients]

    # ── Provider stats ─────────────────────────────────────────────
    provider_stats = Assessment.objects.filter(
        assessment_date__date__range=[start_date, end_date]
    ).values(
        'assessed_by__first_name', 'assessed_by__last_name', 'department'
    ).annotate(
        patient_count=Count('patient', distinct=True),
        assessment_count=Count('id')
    ).order_by('-patient_count')

    provider_data = [
        {
            'name': f"{s['assessed_by__first_name']} {s['assessed_by__last_name']}",
            'department': (s['department'] or 'N/A').title(),
            'patient_count': s['patient_count'],
            'assessment_count': s['assessment_count'],
        }
        for s in provider_stats if s['assessed_by__first_name']
    ]

    # ── Real retention rates ────────────────────────────────────────
    # Definition: of patients who had their first-ever assessment at least N days ago,
    # what % have had a follow-up assessment since?
    def _retention_rate(min_days_since_first):
        cutoff = date.today() - timedelta(days=min_days_since_first)
        # Patients whose first assessment was at least min_days ago
        eligible = Patient.objects.filter(
            is_active=True,
            assessments__assessment_date__date__lte=cutoff
        ).annotate(
            first_assessment=Min('assessments__assessment_date')
        ).filter(first_assessment__date__lte=cutoff).distinct()

        eligible_count = eligible.count()
        if eligible_count == 0:
            return 0
        # Of those, how many have had more than 1 assessment total?
        returned = eligible.annotate(
            total_assessments=Count('assessments')
        ).filter(total_assessments__gt=1).count()
        return round(returned / eligible_count * 100)

    retention_1_month  = _retention_rate(30)
    retention_3_months = _retention_rate(90)
    retention_6_months = _retention_rate(180)
    retention_1_year   = _retention_rate(365)

    # ── Patient table — annotate with visit data ───────────────────
    patients_table = patients.annotate(
        visit_count=Count('assessments'),
        last_visit_date=Max('assessments__assessment_date'),
    ).order_by('-last_visit_date')[:50]

    # ── Cross-department coverage (patients seen in only 1 dept) ───
    patients_only_physio    = patients.filter(
        assessments__department='physiotherapy'
    ).exclude(assessments__department__in=['nutrition', 'general']).distinct()[:50]

    patients_only_nutrition = patients.filter(
        assessments__department='nutrition'
    ).exclude(assessments__department__in=['physiotherapy', 'general']).distinct()[:50]

    patients_only_general   = patients.filter(
        assessments__department='general'
    ).exclude(assessments__department__in=['physiotherapy', 'nutrition']).distinct()[:50]

    context = {
        'date_range': date_range,
        'date_from': date_from,
        'date_to': date_to,
        'age_group': age_group,
        'gender': gender,
        'start_date': start_date,
        'end_date': end_date,
        'total_patients': total_patients,
        'new_patients': new_patients,
        'active_patients': active_patients,
        'average_age': average_age,
        'patients': patients_table,
        'age_labels': json.dumps(age_labels),
        'age_data': json.dumps(age_data),
        'gender_labels': json.dumps(gender_labels),
        'gender_data': json.dumps(gender_data),
        'trend_labels': json.dumps(trend_labels),
        'trend_data': json.dumps(trend_data),
        'insurance_labels': json.dumps(insurance_labels),
        'insurance_data': json.dumps(insurance_data),
        'retention_1_month': retention_1_month,
        'retention_3_months': retention_3_months,
        'retention_6_months': retention_6_months,
        'retention_1_year': retention_1_year,
        'physio_patients': physio_patients,
        'nutrition_patients': nutrition_patients,
        'general_patients': general_patients,
        'dept_labels': json.dumps(dept_labels),
        'dept_data': json.dumps(dept_data),
        'provider_data': provider_data,
        'patients_only_physio': patients_only_physio,
        'patients_only_nutrition': patients_only_nutrition,
        'patients_only_general': patients_only_general,
    }
    return render(request, 'reports/patient_reports.html', context)

@login_required
def financial_reports(request):
    # Get filter parameters
    period = request.GET.get('period', 'this_month')
    service_type = request.GET.get('service_type', '')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Calculate date range based on period
    end_date = date.today()
    
    if period == 'custom' and date_from and date_to:
        # Custom date range
        start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    elif period == 'this_month':
        start_date = end_date.replace(day=1)
    elif period == 'last_month':
        start_date = (end_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        end_date = end_date.replace(day=1) - timedelta(days=1)
    elif period == 'this_quarter':
        quarter = (end_date.month - 1) // 3 + 1
        start_date = date(end_date.year, (quarter - 1) * 3 + 1, 1)
    elif period == 'last_quarter':
        quarter = (end_date.month - 1) // 3 + 1
        if quarter == 1:
            start_date = date(end_date.year - 1, 10, 1)
            end_date = date(end_date.year - 1, 12, 31)
        else:
            start_date = date(end_date.year, (quarter - 2) * 3 + 1, 1)
            end_date = date(end_date.year, (quarter - 1) * 3, 1) - timedelta(days=1)
    elif period == 'this_year':
        start_date = date(end_date.year, 1, 1)
    elif period == 'last_year':
        start_date = date(end_date.year - 1, 1, 1)
        end_date = date(end_date.year - 1, 12, 31)
    else:
        start_date = end_date.replace(day=1)
    
    # Import pharmacy and budget models
    from pharmacy.models import StockMovement, Batch
    from budget.models import Expense, ExpenseCategory
    from django.db.models import F, Q
    
    # Services Revenue (from payments)
    payments_query = Payment.objects.filter(
        payment_date__date__range=[start_date, end_date],
        status='completed'
    )
    
    # Apply service type filter if specified
    if service_type:
        # Filter by service category in related invoices
        payments_query = payments_query.filter(
            invoice__line_items__service__category=service_type
        ).distinct()
    
    services_revenue = payments_query.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Pharmacy Sales Revenue (from stock movements marked as sales)
    pharmacy_sales = StockMovement.objects.filter(
        created_at__date__range=[start_date, end_date],
        movement_type='out',
        reference__icontains='SALE'
    ).annotate(
        revenue=F('quantity') * F('batch__selling_price')
    ).aggregate(total=Sum('revenue'))['total'] or 0
    
    # Total Revenue (Services + Pharmacy Sales)
    total_revenue = services_revenue + pharmacy_sales
    
    # Total Expenses from Budget App (approved and paid expenses)
    total_expenses = Expense.objects.filter(
        expense_date__range=[start_date, end_date],
        status__in=['approved', 'paid']
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Expenses by Category (for breakdown)
    expenses_by_category = Expense.objects.filter(
        expense_date__range=[start_date, end_date],
        status__in=['approved', 'paid']
    ).values('category__name', 'category__color', 'category__icon').annotate(
        total=Sum('amount')
    ).order_by('-total')[:10]  # Top 10 expense categories
    
    # Net Balance (Revenue - Expenses)
    net_balance = total_revenue - total_expenses
    
    # Pharmacy Profit Margin (for reference)
    pharmacy_cogs = StockMovement.objects.filter(
        created_at__date__range=[start_date, end_date],
        movement_type='out',
        reference__icontains='SALE'
    ).annotate(
        cost=F('quantity') * F('batch__cost_price')
    ).aggregate(total=Sum('cost'))['total'] or 0
    
    pharmacy_profit = pharmacy_sales - pharmacy_cogs
    pharmacy_profit_margin = (pharmacy_profit / pharmacy_sales * 100) if pharmacy_sales > 0 else 0
    
    # Apply service type filter to outstanding invoices
    outstanding_query = Invoice.objects.filter(
        status__in=['sent', 'overdue']
    )
    if service_type:
        outstanding_query = outstanding_query.filter(
            line_items__service__category=service_type
        ).distinct()
    outstanding_amount = outstanding_query.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    payments_received = payments_query.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Apply service type filter to invoices
    invoices_query = Invoice.objects.filter(
        created_at__date__range=[start_date, end_date]
    )
    if service_type:
        invoices_query = invoices_query.filter(
            line_items__service__category=service_type
        ).distinct()
    total_invoices = invoices_query.count()
    
    # Calculate collection rate
    collection_rate = (float(payments_received) / float(total_revenue) * 100) if total_revenue > 0 else 0
    
    # Average invoice value
    avg_invoice_value = invoices_query.aggregate(Avg('total_amount'))['total_amount__avg'] or 0
    
    # Revenue trend (entire period, daily breakdown)
    revenue_trend_labels = []
    revenue_trend_data = []
    
    # Calculate number of days in the period
    period_days = (end_date - start_date).days + 1
    
    # Generate daily revenue data for entire period
    for i in range(period_days):
        day = start_date + timedelta(days=i)
        day_query = Payment.objects.filter(
            payment_date__date=day,
            status='completed'
        )
        if service_type:
            day_query = day_query.filter(
                invoice__line_items__service__category=service_type
            ).distinct()
        revenue = day_query.aggregate(Sum('amount'))['amount__sum'] or 0
        revenue_trend_labels.append(day.strftime('%m/%d'))
        revenue_trend_data.append(float(revenue))
    
    # Payment methods
    payment_methods = payments_query.values('payment_method').annotate(total=Sum('amount'))
    
    payment_method_labels = []
    payment_method_data = []
    for item in payment_methods:
        payment_method_labels.append(item['payment_method'])
        payment_method_data.append(float(item['total']))
    
    # Service revenue by category
    service_revenue_query = InvoiceLineItem.objects.filter(
        invoice__created_at__date__range=[start_date, end_date],
        invoice__status__in=['sent', 'paid']
    )
    if service_type:
        service_revenue_query = service_revenue_query.filter(
            service__category=service_type
        )
    service_revenue = service_revenue_query.values('service__category').annotate(
        total_revenue=Sum('total_amount')
    ).order_by('-total_revenue')
    
    service_labels = []
    service_revenue_data = []
    for item in service_revenue:
        category_display = dict(Service.SERVICE_CATEGORIES).get(item['service__category'], item['service__category'])
        service_labels.append(category_display)
        service_revenue_data.append(float(item['total_revenue']))
    
    # Monthly revenue comparison (current vs previous year)
    monthly_labels = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    current_year_data = []
    previous_year_data = []
    
    current_year = date.today().year
    previous_year = current_year - 1
    
    for month in range(1, 13):
        # Current year monthly revenue
        current_month_query = Payment.objects.filter(
            payment_date__year=current_year,
            payment_date__month=month,
            status='completed'
        )
        if service_type:
            current_month_query = current_month_query.filter(
                invoice__line_items__service__category=service_type
            ).distinct()
        current_month_revenue = current_month_query.aggregate(Sum('amount'))['amount__sum'] or 0
        current_year_data.append(float(current_month_revenue))
        
        # Previous year monthly revenue
        previous_month_query = Payment.objects.filter(
            payment_date__year=previous_year,
            payment_date__month=month,
            status='completed'
        )
        if service_type:
            previous_month_query = previous_month_query.filter(
                invoice__line_items__service__category=service_type
            ).distinct()
        previous_month_revenue = previous_month_query.aggregate(Sum('amount'))['amount__sum'] or 0
        previous_year_data.append(float(previous_month_revenue))
    
    # Calculate period comparisons with real data
    # Calculate previous period dates
    if period == 'this_month':
        prev_start = (start_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        prev_end = start_date - timedelta(days=1)
    elif period == 'last_month':
        prev_start = (start_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        prev_end = start_date - timedelta(days=1)
    else:
        # For other periods, calculate equivalent previous period
        period_days = (end_date - start_date).days
        prev_end = start_date - timedelta(days=1)
        prev_start = prev_end - timedelta(days=period_days)
    
    prev_payments_query = Payment.objects.filter(
        payment_date__date__range=[prev_start, prev_end],
        status='completed'
    )
    if service_type:
        prev_payments_query = prev_payments_query.filter(
            invoice__line_items__service__category=service_type
        ).distinct()
    previous_revenue = prev_payments_query.aggregate(Sum('amount'))['amount__sum'] or 0
    
    revenue_change = total_revenue - previous_revenue
    revenue_change_percent = (revenue_change / previous_revenue * 100) if previous_revenue > 0 else 0
    
    # Top performing services
    top_services_query = InvoiceLineItem.objects.filter(
        invoice__created_at__date__range=[start_date, end_date],
        invoice__status__in=['sent', 'paid']
    )
    if service_type:
        top_services_query = top_services_query.filter(
            service__category=service_type
        )
    top_services_data = top_services_query.values('service__name').annotate(
        total_revenue=Sum('total_amount'),
        service_count=Count('id')
    ).order_by('-total_revenue')[:5]
    
    top_services = []
    for item in top_services_data:
        top_services.append({
            'name': item['service__name'] or 'Unknown Service',
            'total_revenue': float(item['total_revenue'] or 0),
            'appointment_count': item['service_count']
        })
    
    # Outstanding invoices by age - calculate real aging data
    from django.utils import timezone
    today = timezone.now().date()
    
    outstanding_0_30 = Invoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__gte=today - timedelta(days=30)
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    outstanding_31_60 = Invoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__range=[today - timedelta(days=60), today - timedelta(days=31)]
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    outstanding_61_90 = Invoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__range=[today - timedelta(days=90), today - timedelta(days=61)]
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    outstanding_90_plus = Invoice.objects.filter(
        status__in=['sent', 'overdue'],
        due_date__lt=today - timedelta(days=90)
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Calculate percentages for aging
    total_outstanding = outstanding_0_30 + outstanding_31_60 + outstanding_61_90 + outstanding_90_plus
    outstanding_0_30_percent = (outstanding_0_30 / total_outstanding * 100) if total_outstanding > 0 else 0
    outstanding_31_60_percent = (outstanding_31_60 / total_outstanding * 100) if total_outstanding > 0 else 0
    outstanding_61_90_percent = (outstanding_61_90 / total_outstanding * 100) if total_outstanding > 0 else 0
    outstanding_90_plus_percent = (outstanding_90_plus / total_outstanding * 100) if total_outstanding > 0 else 0
    
    # Insurance claims data
    from billing.models import InsuranceClaim
    claims_submitted = InsuranceClaim.objects.filter(
        submission_date__range=[start_date, end_date]
    ).count()
    claims_approved = InsuranceClaim.objects.filter(
        submission_date__range=[start_date, end_date],
        status='approved'
    ).count()
    claims_pending = InsuranceClaim.objects.filter(
        submission_date__range=[start_date, end_date],
        status__in=['submitted', 'pending']
    ).count()
    claims_denied = InsuranceClaim.objects.filter(
        submission_date__range=[start_date, end_date],
        status='denied'
    ).count()
    claims_total_amount = InsuranceClaim.objects.filter(
        submission_date__range=[start_date, end_date]
    ).aggregate(Sum('claim_amount'))['claim_amount__sum'] or 0
    claims_reimbursed = InsuranceClaim.objects.filter(
        submission_date__range=[start_date, end_date],
        status__in=['approved', 'paid']
    ).aggregate(Sum('approved_amount'))['approved_amount__sum'] or 0
    
    context = {
        'period': period,
        'service_type': service_type,
        'date_from': date_from,
        'date_to': date_to,
        'total_revenue': total_revenue,
        'services_revenue': services_revenue,
        'pharmacy_sales': pharmacy_sales,
        'total_expenses': total_expenses,
        'expenses_by_category': expenses_by_category,
        'net_balance': net_balance,
        'pharmacy_profit': pharmacy_profit,
        'pharmacy_profit_margin': round(pharmacy_profit_margin, 1),
        'outstanding_amount': outstanding_amount,
        'collection_rate': round(collection_rate, 1),
        'avg_invoice_value': avg_invoice_value,
        'payments_received': payments_received,
        'total_invoices': total_invoices,
        'previous_revenue': previous_revenue,
        'revenue_change': revenue_change,
        'revenue_change_percent': round(revenue_change_percent, 1),
        'previous_invoices': Invoice.objects.filter(
            created_at__date__range=[prev_start, prev_end]
        ).count(),
        'invoice_change': total_invoices - Invoice.objects.filter(
            created_at__date__range=[prev_start, prev_end]
        ).count(),
        'invoice_change_percent': round(((total_invoices - Invoice.objects.filter(
            created_at__date__range=[prev_start, prev_end]
        ).count()) / Invoice.objects.filter(
            created_at__date__range=[prev_start, prev_end]
        ).count() * 100) if Invoice.objects.filter(
            created_at__date__range=[prev_start, prev_end]
        ).count() > 0 else 0, 1),
        'previous_payments': previous_revenue,
        'payment_change': revenue_change,
        'payment_change_percent': round(revenue_change_percent, 1),
        'previous_outstanding': Invoice.objects.filter(
            status__in=['sent', 'overdue'],
            created_at__date__range=[prev_start, prev_end]
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
        'outstanding_change': outstanding_amount - (Invoice.objects.filter(
            status__in=['sent', 'overdue'],
            created_at__date__range=[prev_start, prev_end]
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0),
        'outstanding_change_percent': round(((outstanding_amount - (Invoice.objects.filter(
            status__in=['sent', 'overdue'],
            created_at__date__range=[prev_start, prev_end]
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0)) / (Invoice.objects.filter(
            status__in=['sent', 'overdue'],
            created_at__date__range=[prev_start, prev_end]
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 1) * 100), 1),
        'revenue_trend_labels': json.dumps(revenue_trend_labels),
        'revenue_trend_data': json.dumps(revenue_trend_data),
        'payment_method_labels': json.dumps(payment_method_labels),
        'payment_method_data': json.dumps(payment_method_data),
        'service_labels': json.dumps(service_labels),
        'service_revenue_data': json.dumps(service_revenue_data),
        'monthly_labels': json.dumps(monthly_labels),
        'current_year_data': json.dumps(current_year_data),
        'previous_year_data': json.dumps(previous_year_data),
        'top_services': top_services,
        'outstanding_0_30': outstanding_0_30,
        'outstanding_31_60': outstanding_31_60,
        'outstanding_61_90': outstanding_61_90,
        'outstanding_90_plus': outstanding_90_plus,
        'outstanding_0_30_percent': round(outstanding_0_30_percent, 1),
        'outstanding_31_60_percent': round(outstanding_31_60_percent, 1),
        'outstanding_61_90_percent': round(outstanding_61_90_percent, 1),
        'outstanding_90_plus_percent': round(outstanding_90_plus_percent, 1),
        'claims_submitted': claims_submitted,
        'claims_approved': claims_approved,
        'claims_pending': claims_pending,
        'claims_denied': claims_denied,
        'claims_total_amount': float(claims_total_amount),
        'claims_reimbursed': float(claims_reimbursed),
    }
    return render(request, 'reports/financial_reports.html', context)

@login_required
def appointment_report(request):
    # Date range
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        start_date = date.today().replace(day=1)
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = date.today()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Appointment statistics
    appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date]
    )
    
    total_appointments = appointments.count()
    
    # Status distribution
    status_stats = appointments.values('status').annotate(count=Count('id'))
    
    # Calculate individual status counts
    completed_count = appointments.filter(status='completed').count()
    scheduled_count = appointments.filter(status='scheduled').count()
    cancelled_count = appointments.filter(status='cancelled').count()
    no_show_count = appointments.filter(status='no_show').count()
    
    # Calculate rates
    completion_rate = (completed_count / total_appointments * 100) if total_appointments > 0 else 0
    cancellation_rate = (cancelled_count / total_appointments * 100) if total_appointments > 0 else 0
    no_show_rate = (no_show_count / total_appointments * 100) if total_appointments > 0 else 0
    
    # Service distribution with detailed stats
    service_stats = appointments.values('service__name').annotate(
        count=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        cancelled=Count('id', filter=Q(status='cancelled'))
    ).order_by('-count')
    
    # Provider workload with performance metrics
    provider_stats = appointments.values(
        'provider__first_name', 
        'provider__last_name'
    ).annotate(
        total_count=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        scheduled=Count('id', filter=Q(status='scheduled')),
        cancelled=Count('id', filter=Q(status='cancelled')),
        no_show=Count('id', filter=Q(status='no_show'))
    ).order_by('-total_count')
    
    # Calculate unique patients and providers
    unique_patients = appointments.values('patient').distinct().count()
    unique_providers = appointments.values('provider').distinct().count()
    
    # Average appointments per day
    date_diff = (end_date - start_date).days + 1
    avg_appointments_per_day = total_appointments / date_diff if date_diff > 0 else 0
    
    # Peak hours analysis (if time data available)
    hourly_distribution = []
    for hour in range(8, 18):  # 8 AM to 6 PM
        count = appointments.filter(
            appointment_time__hour=hour
        ).count() if hasattr(appointments.first(), 'appointment_time') else 0
        hourly_distribution.append({
            'hour': f"{hour}:00",
            'count': count
        })
    
    # Daily appointment trend (last 7 days within range)
    daily_labels = []
    daily_data = []
    daily_completed = []
    daily_cancelled = []
    
    current_date = max(start_date, end_date - timedelta(days=6))
    while current_date <= end_date:
        day_appointments = appointments.filter(appointment_date=current_date)
        daily_labels.append(current_date.strftime('%m/%d'))
        daily_data.append(day_appointments.count())
        daily_completed.append(day_appointments.filter(status='completed').count())
        daily_cancelled.append(day_appointments.filter(status='cancelled').count())
        current_date += timedelta(days=1)
    
    # Monthly trend (last 6 months)
    monthly_labels = []
    monthly_data = []
    for i in range(6):
        month_start = (end_date.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        count = Appointment.objects.filter(
            appointment_date__range=[month_start, min(month_end, end_date)]
        ).count()
        monthly_labels.insert(0, month_start.strftime('%b %Y'))
        monthly_data.insert(0, count)
    
    # Department distribution (if applicable)
    dept_stats = appointments.values('service__name').annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    
    # Revenue from completed appointments (if linked to invoices)
    completed_appointments = appointments.filter(status='completed')
    total_revenue = 0
    for apt in completed_appointments:
        if hasattr(apt, 'service') and apt.service:
            total_revenue += apt.service.price
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_appointments': total_appointments,
        'completed_count': completed_count,
        'scheduled_count': scheduled_count,
        'cancelled_count': cancelled_count,
        'no_show_count': no_show_count,
        'completion_rate': round(completion_rate, 1),
        'cancellation_rate': round(cancellation_rate, 1),
        'no_show_rate': round(no_show_rate, 1),
        'unique_patients': unique_patients,
        'unique_providers': unique_providers,
        'avg_appointments_per_day': round(avg_appointments_per_day, 1),
        'total_revenue': total_revenue,
        'status_stats': status_stats,
        'service_stats': service_stats,
        'provider_stats': provider_stats,
        'hourly_distribution': hourly_distribution,
        'daily_labels': json.dumps(daily_labels),
        'daily_data': json.dumps(daily_data),
        'daily_completed': json.dumps(daily_completed),
        'daily_cancelled': json.dumps(daily_cancelled),
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data': json.dumps(monthly_data),
        'dept_stats': dept_stats,
    }
    return render(request, 'reports/appointment_report.html', context)

@login_required
@require_http_methods(["POST"])
def export_report(request):
    """Export report to PDF, Excel, or CSV"""
    try:
        report_type = request.POST.get('report_type')
        export_format = request.POST.get('export_format')
        report_name = request.POST.get('report_name', f'{report_type.title()} Report')
        
        # Get report parameters from POST data
        parameters = {}
        for key, value in request.POST.items():
            if key not in ['report_type', 'export_format', 'report_name', 'csrfmiddlewaretoken']:
                parameters[key] = value
        
        # Generate report data based on type
        if report_type == 'dashboard':
            content_data = generate_dashboard_export_data(parameters)
        elif report_type == 'patient':
            content_data = generate_patient_export_data(parameters)
        elif report_type == 'financial':
            content_data = generate_financial_export_data(parameters)
        elif report_type == 'appointment':
            content_data = generate_appointment_export_data(parameters)
        else:
            return JsonResponse({'error': 'Invalid report type'}, status=400)
        
        # Create export
        response, export_record = create_report_export(
            user=request.user,
            report_type=report_type,
            report_name=report_name,
            export_format=export_format,
            content_data=content_data,
            parameters=parameters
        )
        
        # Log export activity
        audit_mixin = ReportAuditMixin()
        audit_mixin.log_report_activity(
            request=request,
            report_type=report_type,
            report_name=report_name,
            action=f'exported_{export_format}',
            record_count=len(content_data.get('tables', [{}])[0].get('data', [])),
            file_size=export_record.file_size,
            parameters=parameters
        )
        
        return response
        
    except Exception as e:
        # Log error
        audit_mixin = ReportAuditMixin()
        audit_mixin.log_report_activity(
            request=request,
            report_type=request.POST.get('report_type', 'unknown'),
            report_name=request.POST.get('report_name', 'Unknown Report'),
            action=f'exported_{request.POST.get("export_format", "unknown")}',
            success=False,
            error_message=str(e)
        )
        return JsonResponse({'error': str(e)}, status=500)

def generate_dashboard_export_data(parameters):
    """Generate export data for dashboard report"""
    # Parse date parameters
    start_date = datetime.strptime(parameters.get('start_date', date.today().replace(day=1).isoformat()), '%Y-%m-%d').date()
    end_date = datetime.strptime(parameters.get('end_date', date.today().isoformat()), '%Y-%m-%d').date()
    
    # Get dashboard data
    total_patients = Patient.objects.filter(is_active=True).count()
    new_patients = Patient.objects.filter(registration_date__range=[start_date, end_date]).count()
    total_appointments = Appointment.objects.filter(appointment_date__range=[start_date, end_date]).count()
    completed_appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date], status='completed'
    ).count()
    total_revenue = Payment.objects.filter(
        payment_date__date__range=[start_date, end_date], status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Popular services data
    popular_services = Service.objects.annotate(
        appointment_count=Count('appointment')
    ).order_by('-appointment_count')[:10]
    
    services_data = []
    for service in popular_services:
        services_data.append([
            service.name,
            service.category,
            f"UGX {service.price:,.0f}",
            service.appointment_count
        ])
    
    return {
        'summary_stats': {
            'Total Patients': total_patients,
            'New Patients': new_patients,
            'Total Appointments': total_appointments,
            'Completed Appointments': completed_appointments,
            'Total Revenue': f"UGX {total_revenue:,.0f}",
            'Report Period': f"{start_date} to {end_date}"
        },
        'tables': [
            {
                'title': 'Popular Services',
                'headers': ['Service Name', 'Category', 'Price', 'Appointments'],
                'data': services_data
            }
        ]
    }

def generate_patient_export_data(parameters):
    """Generate export data for patient report"""
    # Get filter parameters
    date_range = parameters.get('date_range', 'last_30_days')
    gender = parameters.get('gender', '')
    
    # Calculate date range
    end_date = date.today()
    if date_range == 'last_30_days':
        start_date = end_date - timedelta(days=30)
    elif date_range == 'last_3_months':
        start_date = end_date - timedelta(days=90)
    elif date_range == 'last_6_months':
        start_date = end_date - timedelta(days=180)
    elif date_range == 'last_year':
        start_date = end_date - timedelta(days=365)
    else:
        start_date = end_date - timedelta(days=30)
    
    # Base patient queryset
    patients = Patient.objects.filter(is_active=True)
    if gender:
        patients = patients.filter(gender=gender)
    
    # Patient data for export
    patient_data = []
    for patient in patients[:100]:  # Limit for export
        try:
            age = patient.get_age() if hasattr(patient, 'get_age') else None
            age_display = str(int(age)) if age and isinstance(age, (int, float)) else 'N/A'
        except (ValueError, TypeError, AttributeError):
            age_display = 'N/A'
            
        patient_data.append([
            f"{patient.first_name} {patient.last_name}",
            patient.email,
            patient.phone,
            patient.gender,
            age_display,
            patient.registration_date.strftime('%Y-%m-%d'),
            patient.insurance_provider or 'None'
        ])
    
    # Calculate average age safely
    ages = []
    for p in patients:
        try:
            age = p.get_age() if hasattr(p, 'get_age') else None
            if age and isinstance(age, (int, float)):
                ages.append(int(age))
        except (ValueError, TypeError, AttributeError):
            continue
    
    avg_age = round(sum(ages) / len(ages), 1) if ages else 0
    
    return {
        'summary_stats': {
            'Total Patients': patients.count(),
            'New Patients (Period)': patients.filter(registration_date__range=[start_date, end_date]).count(),
            'Average Age': avg_age,
            'Report Period': f"{start_date} to {end_date}"
        },
        'tables': [
            {
                'title': 'Patient List',
                'headers': ['Name', 'Email', 'Phone', 'Gender', 'Age', 'Registration Date', 'Insurance'],
                'data': patient_data
            }
        ]
    }

def generate_financial_export_data(parameters):
    """Generate export data for financial report"""
    period = parameters.get('period', 'this_month')
    
    # Calculate date range based on period
    end_date = date.today()
    if period == 'this_month':
        start_date = end_date.replace(day=1)
    elif period == 'last_month':
        start_date = (end_date.replace(day=1) - timedelta(days=1)).replace(day=1)
        end_date = end_date.replace(day=1) - timedelta(days=1)
    elif period == 'this_quarter':
        quarter = (end_date.month - 1) // 3 + 1
        start_date = date(end_date.year, (quarter - 1) * 3 + 1, 1)
    elif period == 'this_year':
        start_date = date(end_date.year, 1, 1)
    else:
        start_date = end_date.replace(day=1)
    
    # Financial metrics
    total_revenue = Payment.objects.filter(
        payment_date__date__range=[start_date, end_date], status='completed'
    ).aggregate(Sum('amount'))['amount__sum'] or 0
    
    outstanding_amount = Invoice.objects.filter(
        status__in=['sent', 'overdue']
    ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Recent invoices data
    recent_invoices = Invoice.objects.filter(
        created_at__date__range=[start_date, end_date]
    ).order_by('-created_at')[:50]
    
    invoice_data = []
    for invoice in recent_invoices:
        invoice_data.append([
            invoice.invoice_number,
            f"{invoice.patient.first_name} {invoice.patient.last_name}" if invoice.patient else 'N/A',
            f"UGX {invoice.total_amount:,.0f}",
            invoice.status.title(),
            invoice.created_at.strftime('%Y-%m-%d'),
            invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else 'N/A'
        ])
    
    return {
        'summary_stats': {
            'Total Revenue': f"UGX {total_revenue:,.0f}",
            'Outstanding Amount': f"UGX {outstanding_amount:,.0f}",
            'Total Invoices': recent_invoices.count(),
            'Report Period': f"{start_date} to {end_date}"
        },
        'tables': [
            {
                'title': 'Recent Invoices',
                'headers': ['Invoice #', 'Patient', 'Amount', 'Status', 'Created', 'Due Date'],
                'data': invoice_data
            }
        ]
    }

def generate_appointment_export_data(parameters):
    """Generate export data for appointment report"""
    start_date = datetime.strptime(parameters.get('start_date', date.today().replace(day=1).isoformat()), '%Y-%m-%d').date()
    end_date = datetime.strptime(parameters.get('end_date', date.today().isoformat()), '%Y-%m-%d').date()
    
    # Appointment data
    appointments = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date]
    ).order_by('-appointment_date')[:100]
    
    appointment_data = []
    for appointment in appointments:
        appointment_data.append([
            appointment.appointment_date.strftime('%Y-%m-%d %H:%M'),
            f"{appointment.patient.first_name} {appointment.patient.last_name}" if appointment.patient else 'N/A',
            f"{appointment.provider.first_name} {appointment.provider.last_name}" if appointment.provider else 'N/A',
            appointment.service.name if appointment.service else 'N/A',
            appointment.status.title(),
            appointment.notes[:100] + '...' if appointment.notes and len(appointment.notes) > 100 else appointment.notes or ''
        ])
    
    return {
        'summary_stats': {
            'Total Appointments': appointments.count(),
            'Completed': appointments.filter(status='completed').count(),
            'Cancelled': appointments.filter(status='cancelled').count(),
            'No Show': appointments.filter(status='no_show').count(),
            'Report Period': f"{start_date} to {end_date}"
        },
        'tables': [
            {
                'title': 'Appointments',
                'headers': ['Date/Time', 'Patient', 'Provider', 'Service', 'Status', 'Notes'],
                'data': appointment_data
            }
        ]
    }

@login_required
def audit_log(request):
    """View audit log for reports"""
    # Get filter parameters
    user_filter = request.GET.get('user')
    report_type_filter = request.GET.get('report_type')
    action_filter = request.GET.get('action')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset
    audit_logs = ReportAuditLog.objects.all()
    
    # Apply filters
    if user_filter:
        audit_logs = audit_logs.filter(user__username__icontains=user_filter)
    if report_type_filter:
        audit_logs = audit_logs.filter(report_type=report_type_filter)
    if action_filter:
        audit_logs = audit_logs.filter(action=action_filter)
    if date_from:
        audit_logs = audit_logs.filter(timestamp__date__gte=date_from)
    if date_to:
        audit_logs = audit_logs.filter(timestamp__date__lte=date_to)
    
    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(audit_logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    users = User.objects.filter(report_activities__isnull=False).distinct()
    report_types = ReportAuditLog.objects.values_list('report_type', flat=True).distinct()
    actions = ReportAuditLog.objects.values_list('action', flat=True).distinct()
    
    context = {
        'page_obj': page_obj,
        'users': users,
        'report_types': report_types,
        'actions': actions,
        'filters': {
            'user': user_filter,
            'report_type': report_type_filter,
            'action': action_filter,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    return render(request, 'reports/audit_log.html', context)

@login_required
def report_performance(request):
    """View report performance metrics with comprehensive analytics"""
    metrics = get_report_performance_metrics()
    
    # Get detailed performance data
    from django.db.models import Avg, Max, Min, Sum
    
    # Date range for analysis
    end_date = date.today()
    start_date = end_date - timedelta(days=30)
    
    # Total reports generated
    total_reports = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date]
    ).count()
    
    # Success rate
    successful_reports = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date],
        success=True
    ).count()
    success_rate = (successful_reports / total_reports * 100) if total_reports > 0 else 0
    
    # Execution time statistics by report type
    execution_stats = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date],
        success=True
    ).values('report_type').annotate(
        avg_time=Avg('execution_time'),
        max_time=Max('execution_time'),
        min_time=Min('execution_time'),
        total_count=Count('id'),
        total_records=Sum('record_count')
    ).order_by('-avg_time')
    
    # Most used reports
    popular_reports = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date]
    ).values('report_type', 'report_name').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # User activity
    user_activity = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date]
    ).values('user__first_name', 'user__last_name').annotate(
        report_count=Count('id'),
        avg_execution_time=Avg('execution_time')
    ).order_by('-report_count')[:10]
    
    # Action distribution
    action_stats = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date]
    ).values('action').annotate(
        count=Count('id')
    ).order_by('-count')
    
    action_labels = []
    action_data = []
    for stat in action_stats:
        action_labels.append(stat['action'].title())
        action_data.append(stat['count'])
    
    # Daily report generation trend (last 30 days)
    daily_stats = []
    daily_success = []
    daily_failures = []
    
    for i in range(30):
        day = end_date - timedelta(days=i)
        total = ReportAuditLog.objects.filter(
            timestamp__date=day
        ).count()
        success = ReportAuditLog.objects.filter(
            timestamp__date=day,
            success=True
        ).count()
        failed = total - success
        
        daily_stats.insert(0, {
            'date': day.strftime('%m/%d'),
            'count': total,
            'success': success,
            'failed': failed
        })
        daily_success.insert(0, success)
        daily_failures.insert(0, failed)
    
    # Hourly distribution (peak usage times)
    hourly_stats = []
    for hour in range(24):
        count = ReportAuditLog.objects.filter(
            timestamp__date__range=[start_date, end_date],
            timestamp__hour=hour
        ).count()
        hourly_stats.append({
            'hour': f"{hour}:00",
            'count': count
        })
    
    # Performance recommendations
    recommendations = []
    
    # Check for slow reports
    slow_reports = execution_stats.filter(avg_time__gt=5)
    if slow_reports.exists():
        recommendations.append({
            'type': 'warning',
            'message': f"{slow_reports.count()} report type(s) have average execution time > 5 seconds. Consider optimization."
        })
    
    # Check failure rate
    if success_rate < 95:
        recommendations.append({
            'type': 'danger',
            'message': f"Success rate is {success_rate:.1f}%. Investigate failed reports."
        })
    
    # Check for high usage
    if total_reports > 1000:
        recommendations.append({
            'type': 'info',
            'message': f"{total_reports} reports generated in last 30 days. Consider implementing caching."
        })
    
    # Average response time
    avg_response_time = ReportAuditLog.objects.filter(
        timestamp__date__range=[start_date, end_date],
        success=True
    ).aggregate(Avg('execution_time'))['execution_time__avg'] or 0
    
    # Fastest and slowest reports
    fastest_report = execution_stats.order_by('avg_time').first()
    slowest_report = execution_stats.order_by('-avg_time').first()
    
    context = {
        'metrics': metrics,
        'total_reports': total_reports,
        'successful_reports': successful_reports,
        'success_rate': round(success_rate, 1),
        'avg_response_time': round(avg_response_time, 3),
        'fastest_report': fastest_report,
        'slowest_report': slowest_report,
        'execution_stats': execution_stats,
        'popular_reports': popular_reports,
        'user_activity': user_activity,
        'action_stats': action_stats,
        'action_labels': json.dumps(action_labels),
        'action_data': json.dumps(action_data),
        'daily_stats': daily_stats,
        'daily_labels': json.dumps([stat['date'] for stat in daily_stats]),
        'daily_data': json.dumps([stat['count'] for stat in daily_stats]),
        'daily_success': json.dumps(daily_success),
        'daily_failures': json.dumps(daily_failures),
        'hourly_stats': hourly_stats,
        'recommendations': recommendations,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'reports/performance.html', context)

@login_required
def physiotherapy_reports(request):
    """Physiotherapy department specific reports"""
    from patients.models import Assessment
    
    # Get filter parameters
    start_date = request.GET.get('start_date', date.today().replace(day=1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Physiotherapy assessments
    physio_assessments = Assessment.objects.filter(
        department='physiotherapy',
        assessment_date__range=[start_date, end_date]
    ).select_related('patient', 'assessed_by')
    
    # Statistics
    total_assessments = physio_assessments.count()
    first_visits = physio_assessments.filter(assessment_type='first_visit').count()
    follow_ups = physio_assessments.filter(assessment_type='follow_up').count()
    
    # Common diagnoses
    diagnoses = {}
    for assessment in physio_assessments:
        if assessment.diagnosis:
            diagnosis = assessment.diagnosis.strip()
            diagnoses[diagnosis] = diagnoses.get(diagnosis, 0) + 1
    
    common_diagnoses = sorted(diagnoses.items(), key=lambda x: x[1], reverse=True)[:10]
    
    # Treatment outcomes
    follow_up_required = physio_assessments.filter(follow_up_required=True).count()
    
    # Therapist performance
    therapist_stats = physio_assessments.values('assessed_by__first_name', 'assessed_by__last_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Monthly trend
    monthly_labels = []
    monthly_data = []
    for i in range(6):
        month_start = (end_date.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        count = Assessment.objects.filter(
            department='physiotherapy',
            assessment_date__range=[month_start, month_end]
        ).count()
        
        monthly_labels.insert(0, month_start.strftime('%b %Y'))
        monthly_data.insert(0, count)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_assessments': total_assessments,
        'first_visits': first_visits,
        'follow_ups': follow_ups,
        'common_diagnoses': common_diagnoses,
        'follow_up_required': follow_up_required,
        'therapist_stats': therapist_stats,
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data': json.dumps(monthly_data),
        'recent_assessments': physio_assessments.order_by('-assessment_date')[:20],
    }
    return render(request, 'reports/physiotherapy_reports.html', context)

@login_required
def nutrition_reports(request):
    """Nutrition department specific reports"""
    from patients.models import Assessment
    
    # Get filter parameters
    start_date = request.GET.get('start_date', date.today().replace(day=1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Nutrition assessments
    nutrition_assessments = Assessment.objects.filter(
        department='nutrition',
        assessment_date__range=[start_date, end_date]
    ).select_related('patient', 'assessed_by')
    
    # Statistics
    total_assessments = nutrition_assessments.count()
    first_visits = nutrition_assessments.filter(assessment_type='first_visit').count()
    follow_ups = nutrition_assessments.filter(assessment_type='follow_up').count()
    
    # Common diagnoses/conditions
    diagnoses = {}
    for assessment in nutrition_assessments:
        if assessment.diagnosis:
            diagnosis = assessment.diagnosis.strip()
            diagnoses[diagnosis] = diagnoses.get(diagnosis, 0) + 1
    
    common_conditions = sorted(diagnoses.items(), key=lambda x: x[1], reverse=True)[:10]
    
    # Follow-up tracking
    follow_up_required = nutrition_assessments.filter(follow_up_required=True).count()
    
    # Nutritionist performance
    nutritionist_stats = nutrition_assessments.values('assessed_by__first_name', 'assessed_by__last_name').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Monthly trend
    monthly_labels = []
    monthly_data = []
    for i in range(6):
        month_start = (end_date.replace(day=1) - timedelta(days=i*30)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        count = Assessment.objects.filter(
            department='nutrition',
            assessment_date__range=[month_start, month_end]
        ).count()
        
        monthly_labels.insert(0, month_start.strftime('%b %Y'))
        monthly_data.insert(0, count)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_assessments': total_assessments,
        'first_visits': first_visits,
        'follow_ups': follow_ups,
        'common_conditions': common_conditions,
        'follow_up_required': follow_up_required,
        'nutritionist_stats': nutritionist_stats,
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data': json.dumps(monthly_data),
        'recent_assessments': nutrition_assessments.order_by('-assessment_date')[:20],
    }
    return render(request, 'reports/nutrition_reports.html', context)

@login_required
def statistics(request):
    """Comprehensive period statistics: daily, weekly, monthly, annual, or custom date range"""
    from patients.models import Assessment
    from calendar import monthrange

    period = request.GET.get('period', 'monthly')
    today = date.today()

    # ── Custom date range takes priority over period shortcuts ──────
    custom_start = request.GET.get('start_date', '').strip()
    custom_end   = request.GET.get('end_date', '').strip()

    if custom_start and custom_end:
        try:
            start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
            end_date   = datetime.strptime(custom_end,   '%Y-%m-%d').date()
            if start_date > end_date:
                start_date, end_date = end_date, start_date
            span_days  = (end_date - start_date).days + 1
            prev_start = start_date - timedelta(days=span_days)
            prev_end   = start_date - timedelta(days=1)
            period     = 'custom'
            period_label = f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}"
            trend_days = span_days
        except ValueError:
            custom_start = custom_end = ''

    if not (custom_start and custom_end):
        # ── Period boundaries ──────────────────────────────────────
        if period == 'daily':
            start_date = today
            end_date = today
            prev_start = today - timedelta(days=1)
            prev_end = today - timedelta(days=1)
            trend_days = 14
            period_label = today.strftime('%A, %B %d %Y')
        elif period == 'weekly':
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
            prev_start = start_date - timedelta(days=7)
            prev_end = end_date - timedelta(days=7)
            trend_days = 12
            period_label = f"Week of {start_date.strftime('%b %d')} – {end_date.strftime('%b %d, %Y')}"
        elif period == 'annual':
            start_date = date(today.year, 1, 1)
            end_date = date(today.year, 12, 31)
            prev_start = date(today.year - 1, 1, 1)
            prev_end = date(today.year - 1, 12, 31)
            trend_days = None
            period_label = f"Year {today.year}"
        else:  # monthly (default)
            period = 'monthly'
            start_date = today.replace(day=1)
            _, last_day = monthrange(today.year, today.month)
            end_date = today.replace(day=last_day)
            prev_month_end = start_date - timedelta(days=1)
            prev_start = prev_month_end.replace(day=1)
            prev_end = prev_month_end
            trend_days = 12
            period_label = today.strftime('%B %Y')

    # ── Core stats helpers ─────────────────────────────────────────
    def _patients_in(s, e):
        return Patient.objects.filter(registration_date__date__range=[s, e]).count()

    def _assessments_in(s, e):
        return Assessment.objects.filter(assessment_date__date__range=[s, e]).count()

    def _appointments_in(s, e, status=None):
        q = Appointment.objects.filter(appointment_date__range=[s, e])
        if status:
            q = q.filter(status=status)
        return q.count()

    def _revenue_in(s, e):
        return float(Payment.objects.filter(
            payment_date__date__range=[s, e], status='completed'
        ).aggregate(total=Sum('amount'))['total'] or 0)

    # ── Current period ─────────────────────────────────────────────
    new_patients       = _patients_in(start_date, end_date)
    total_assessments  = _assessments_in(start_date, end_date)
    total_appts        = _appointments_in(start_date, end_date)
    completed_appts    = _appointments_in(start_date, end_date, 'completed')
    cancelled_appts    = _appointments_in(start_date, end_date, 'cancelled')
    no_show_appts      = _appointments_in(start_date, end_date, 'no_show')
    revenue            = _revenue_in(start_date, end_date)

    # ── Previous period (for % change) ─────────────────────────────
    prev_new_patients  = _patients_in(prev_start, prev_end)
    prev_assessments   = _assessments_in(prev_start, prev_end)
    prev_revenue       = _revenue_in(prev_start, prev_end)

    def _pct(curr, prev):
        if prev == 0:
            return None
        return round((curr - prev) / prev * 100, 1)

    patient_change    = _pct(new_patients, prev_new_patients)
    assessment_change = _pct(total_assessments, prev_assessments)
    revenue_change    = _pct(revenue, prev_revenue)

    # ── Department breakdown ───────────────────────────────────────
    dept_physio   = Assessment.objects.filter(department='physiotherapy', assessment_date__date__range=[start_date, end_date]).count()
    dept_nutrition = Assessment.objects.filter(department='nutrition',    assessment_date__date__range=[start_date, end_date]).count()
    dept_general  = Assessment.objects.filter(department='general',       assessment_date__date__range=[start_date, end_date]).count()

    # ── Return patient stats ───────────────────────────────────────
    total_active = Patient.objects.filter(is_active=True).count()
    returning_patients = Patient.objects.filter(
        is_active=True,
        assessments__isnull=False
    ).annotate(visit_count=Count('assessments')).filter(visit_count__gt=1).count()
    new_first_visit = Assessment.objects.filter(
        assessment_date__date__range=[start_date, end_date],
        assessment_type='first_visit'
    ).count()
    follow_up_count = Assessment.objects.filter(
        assessment_date__date__range=[start_date, end_date],
        assessment_type='follow_up'
    ).count()
    return_rate = round(follow_up_count / total_assessments * 100, 1) if total_assessments > 0 else 0

    # Top returning patients — filtered by selected date range
    top_returning = Patient.objects.filter(
        is_active=True,
        assessments__assessment_date__date__range=[start_date, end_date]
    ).annotate(
        visit_count=Count('assessments', filter=Q(
            assessments__assessment_date__date__range=[start_date, end_date]
        ))
    ).filter(visit_count__gt=0).order_by('-visit_count')[:10]

    # ── Trend chart data ───────────────────────────────────────────
    trend_labels   = []
    trend_patients = []
    trend_visits   = []
    trend_revenue  = []

    if period == 'daily':
        for i in range(trend_days - 1, -1, -1):
            d = today - timedelta(days=i)
            trend_labels.append(d.strftime('%m/%d'))
            trend_patients.append(_patients_in(d, d))
            trend_visits.append(_assessments_in(d, d))
            trend_revenue.append(_revenue_in(d, d))

    elif period == 'weekly':
        for i in range(trend_days - 1, -1, -1):
            wk_start = start_date - timedelta(weeks=i)
            wk_end   = wk_start + timedelta(days=6)
            trend_labels.append(wk_start.strftime('%m/%d'))
            trend_patients.append(_patients_in(wk_start, wk_end))
            trend_visits.append(_assessments_in(wk_start, wk_end))
            trend_revenue.append(_revenue_in(wk_start, wk_end))

    elif period == 'monthly':
        for i in range(trend_days - 1, -1, -1):
            m_ref = today.month - i
            y_ref = today.year
            while m_ref <= 0:
                m_ref += 12
                y_ref -= 1
            _, ml = monthrange(y_ref, m_ref)
            ms = date(y_ref, m_ref, 1)
            me = date(y_ref, m_ref, ml)
            trend_labels.append(ms.strftime('%b %Y'))
            trend_patients.append(_patients_in(ms, me))
            trend_visits.append(_assessments_in(ms, me))
            trend_revenue.append(_revenue_in(ms, me))

    elif period == 'annual':
        for yr in range(today.year - 4, today.year + 1):
            ys = date(yr, 1, 1)
            ye = date(yr, 12, 31)
            trend_labels.append(str(yr))
            trend_patients.append(_patients_in(ys, ye))
            trend_visits.append(_assessments_in(ys, ye))
            trend_revenue.append(_revenue_in(ys, ye))

    elif period == 'custom':
        span = (end_date - start_date).days
        if span <= 31:
            # Daily granularity
            d = start_date
            while d <= end_date:
                trend_labels.append(d.strftime('%m/%d'))
                trend_patients.append(_patients_in(d, d))
                trend_visits.append(_assessments_in(d, d))
                trend_revenue.append(_revenue_in(d, d))
                d += timedelta(days=1)
        elif span <= 90:
            # Weekly granularity
            d = start_date
            while d <= end_date:
                wk_end = min(d + timedelta(days=6), end_date)
                trend_labels.append(d.strftime('%m/%d'))
                trend_patients.append(_patients_in(d, wk_end))
                trend_visits.append(_assessments_in(d, wk_end))
                trend_revenue.append(_revenue_in(d, wk_end))
                d += timedelta(days=7)
        else:
            # Monthly granularity
            cur = start_date.replace(day=1)
            while cur <= end_date:
                _, ml = monthrange(cur.year, cur.month)
                me = date(cur.year, cur.month, ml)
                trend_labels.append(cur.strftime('%b %Y'))
                trend_patients.append(_patients_in(cur, me))
                trend_visits.append(_assessments_in(cur, me))
                trend_revenue.append(_revenue_in(cur, me))
                if cur.month == 12:
                    cur = date(cur.year + 1, 1, 1)
                else:
                    cur = date(cur.year, cur.month + 1, 1)

    # ── New vs Returning breakdown — follows selected date range ────
    retention_labels = []
    retention_new    = []
    retention_return = []

    def _nv_fup(s, e):
        nv  = Assessment.objects.filter(assessment_date__date__range=[s, e], assessment_type='first_visit').count()
        fup = Assessment.objects.filter(assessment_date__date__range=[s, e], assessment_type='follow_up').count()
        return nv, fup

    span_days_ret = (end_date - start_date).days

    if period == 'daily' or span_days_ret <= 31:
        # Daily buckets
        d = start_date
        while d <= end_date:
            nv, fup = _nv_fup(d, d)
            retention_labels.append(d.strftime('%m/%d'))
            retention_new.append(nv)
            retention_return.append(fup)
            d += timedelta(days=1)
    elif period == 'weekly' or span_days_ret <= 90:
        # Weekly buckets
        d = start_date
        while d <= end_date:
            wk_end = min(d + timedelta(days=6), end_date)
            nv, fup = _nv_fup(d, wk_end)
            retention_labels.append(d.strftime('%m/%d'))
            retention_new.append(nv)
            retention_return.append(fup)
            d += timedelta(days=7)
    elif period == 'annual':
        # Each month of the year
        for m in range(1, 13):
            _, ml = monthrange(start_date.year, m)
            ms = date(start_date.year, m, 1)
            me = date(start_date.year, m, ml)
            nv, fup = _nv_fup(ms, me)
            retention_labels.append(ms.strftime('%b'))
            retention_new.append(nv)
            retention_return.append(fup)
    else:
        # Monthly buckets (monthly period or custom > 90 days)
        cur = start_date.replace(day=1)
        while cur <= end_date:
            _, ml = monthrange(cur.year, cur.month)
            me = date(cur.year, cur.month, ml)
            bucket_end = min(me, end_date)
            nv, fup = _nv_fup(cur, bucket_end)
            retention_labels.append(cur.strftime('%b %Y'))
            retention_new.append(nv)
            retention_return.append(fup)
            if cur.month == 12:
                cur = date(cur.year + 1, 1, 1)
            else:
                cur = date(cur.year, cur.month + 1, 1)

    # Derive a human-readable chart subtitle for the template
    retention_subtitle = f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}"

    context = {
        'period': period,
        'period_label': period_label,
        'start_date': start_date,
        'end_date': end_date,
        'start_date_iso': start_date.isoformat(),
        'end_date_iso': end_date.isoformat(),
        # Current period metrics
        'new_patients': new_patients,
        'total_assessments': total_assessments,
        'total_appts': total_appts,
        'completed_appts': completed_appts,
        'cancelled_appts': cancelled_appts,
        'no_show_appts': no_show_appts,
        'revenue': revenue,
        # Changes
        'patient_change': patient_change,
        'assessment_change': assessment_change,
        'revenue_change': revenue_change,
        # Department
        'dept_physio': dept_physio,
        'dept_nutrition': dept_nutrition,
        'dept_general': dept_general,
        # Return stats
        'total_active': total_active,
        'returning_patients': returning_patients,
        'new_first_visit': new_first_visit,
        'follow_up_count': follow_up_count,
        'return_rate': return_rate,
        'top_returning': top_returning,
        # Charts
        'trend_labels': json.dumps(trend_labels),
        'trend_patients': json.dumps(trend_patients),
        'trend_visits': json.dumps(trend_visits),
        'trend_revenue': json.dumps(trend_revenue),
        'retention_labels': json.dumps(retention_labels),
        'retention_new': json.dumps(retention_new),
        'retention_return': json.dumps(retention_return),
        'retention_subtitle': retention_subtitle,
    }
    return render(request, 'reports/statistics.html', context)


@login_required
def export_statistics_excel(request):
    """Export statistics as a multi-sheet Excel file matching the selected date range."""
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from patients.models import Assessment
    from calendar import monthrange

    # ── Resolve date range (same logic as statistics view) ──────────
    today = date.today()
    period = request.GET.get('period', 'monthly')
    custom_start = request.GET.get('start_date', '').strip()
    custom_end   = request.GET.get('end_date', '').strip()

    if custom_start and custom_end:
        try:
            start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
            end_date   = datetime.strptime(custom_end,   '%Y-%m-%d').date()
            if start_date > end_date:
                start_date, end_date = end_date, start_date
            period_label = f"{start_date.strftime('%b %d, %Y')} – {end_date.strftime('%b %d, %Y')}"
            period = 'custom'
        except ValueError:
            custom_start = custom_end = ''

    if not (custom_start and custom_end):
        if period == 'daily':
            start_date = end_date = today
            period_label = today.strftime('%A, %B %d %Y')
        elif period == 'weekly':
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
            period_label = f"Week {start_date.strftime('%b %d')} – {end_date.strftime('%b %d, %Y')}"
        elif period == 'annual':
            start_date = date(today.year, 1, 1)
            end_date   = date(today.year, 12, 31)
            period_label = f"Year {today.year}"
        else:
            period = 'monthly'
            start_date = today.replace(day=1)
            _, ld = monthrange(today.year, today.month)
            end_date = today.replace(day=ld)
            period_label = today.strftime('%B %Y')

    span_days = (end_date - start_date).days

    # ── Helper queries ───────────────────────────────────────────────
    def _patients_in(s, e):
        return Patient.objects.filter(registration_date__date__range=[s, e]).count()

    def _assessments_in(s, e):
        return Assessment.objects.filter(assessment_date__date__range=[s, e]).count()

    def _appts_in(s, e, status=None):
        q = Appointment.objects.filter(appointment_date__range=[s, e])
        if status:
            q = q.filter(status=status)
        return q.count()

    def _revenue_in(s, e):
        return float(Payment.objects.filter(
            payment_date__date__range=[s, e], status='completed'
        ).aggregate(t=Sum('amount'))['t'] or 0)

    def _nv_fup(s, e):
        nv  = Assessment.objects.filter(assessment_date__date__range=[s, e], assessment_type='first_visit').count()
        fup = Assessment.objects.filter(assessment_date__date__range=[s, e], assessment_type='follow_up').count()
        return nv, fup

    # ── Shared style helpers ─────────────────────────────────────────
    HDR_FILL  = PatternFill('solid', fgColor='1B5E96')
    HDR_FONT  = Font(bold=True, color='FFFFFF', size=11)
    SUB_FILL  = PatternFill('solid', fgColor='E8F0FB')
    SUB_FONT  = Font(bold=True, size=10)
    BOLD_FONT = Font(bold=True)
    CENTRE    = Alignment(horizontal='center', vertical='center')
    WRAP      = Alignment(wrap_text=True)
    THIN      = Side(border_style='thin', color='CCCCCC')
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    NUM_FMT   = '#,##0'
    MONEY_FMT = '#,##0.00'

    def set_header_row(ws, row, cols):
        for col, title in enumerate(cols, 1):
            c = ws.cell(row=row, column=col, value=title)
            c.fill  = HDR_FILL
            c.font  = HDR_FONT
            c.alignment = CENTRE
            c.border = BORDER

    def auto_width(ws, min_w=12, max_w=40):
        for col in ws.columns:
            best = min_w
            for cell in col:
                try:
                    best = max(best, min(len(str(cell.value or '')), max_w))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = best + 2

    def apply_border_row(ws, row, ncols):
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).border = BORDER

    # ── Workbook ─────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)          # remove default blank sheet

    # ════════════════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ════════════════════════════════════════════════════════════════
    ws = wb.create_sheet('Summary')
    ws.sheet_properties.tabColor = '1B5E96'

    # Title block
    ws.merge_cells('A1:D1')
    t = ws['A1']
    t.value = f'Clinic Statistics Summary — {period_label}'
    t.font  = Font(bold=True, size=14, color='1B5E96')
    t.alignment = CENTRE

    ws.merge_cells('A2:D2')
    ws['A2'].value = f'Date range: {start_date.strftime("%B %d, %Y")}  to  {end_date.strftime("%B %d, %Y")}'
    ws['A2'].alignment = CENTRE
    ws['A2'].font = Font(italic=True, color='666666')

    # Section: Patient Metrics
    ws['A4'].value = 'PATIENT METRICS'
    ws['A4'].fill  = SUB_FILL
    ws['A4'].font  = SUB_FONT
    ws.merge_cells('A4:D4')

    metrics = [
        ('New Patients Registered',   _patients_in(start_date, end_date)),
        ('Total Assessments / Visits', _assessments_in(start_date, end_date)),
        ('First Visits',               Assessment.objects.filter(assessment_date__date__range=[start_date, end_date], assessment_type='first_visit').count()),
        ('Follow-up Visits',           Assessment.objects.filter(assessment_date__date__range=[start_date, end_date], assessment_type='follow_up').count()),
    ]
    row = 5
    set_header_row(ws, row, ['Metric', 'Value', '', ''])
    row += 1
    for label, val in metrics:
        ws.cell(row=row, column=1, value=label).border = BORDER
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = NUM_FMT
        c.border = BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='APPOINTMENT METRICS').fill = SUB_FILL
    ws.cell(row=row, column=1).font = SUB_FONT
    ws.merge_cells(f'A{row}:D{row}')
    row += 1

    set_header_row(ws, row, ['Status', 'Count', '', ''])
    row += 1
    appt_rows = [
        ('Total Scheduled',  _appts_in(start_date, end_date)),
        ('Completed',        _appts_in(start_date, end_date, 'completed')),
        ('Cancelled',        _appts_in(start_date, end_date, 'cancelled')),
        ('No-Show',          _appts_in(start_date, end_date, 'no_show')),
    ]
    for label, val in appt_rows:
        ws.cell(row=row, column=1, value=label).border = BORDER
        c = ws.cell(row=row, column=2, value=val)
        c.number_format = NUM_FMT
        c.border = BORDER
        row += 1

    row += 1
    ws.cell(row=row, column=1, value='FINANCIAL METRICS').fill = SUB_FILL
    ws.cell(row=row, column=1).font = SUB_FONT
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    set_header_row(ws, row, ['Metric', 'Amount (UGX)', '', ''])
    row += 1
    ws.cell(row=row, column=1, value='Revenue Collected').border = BORDER
    c = ws.cell(row=row, column=2, value=_revenue_in(start_date, end_date))
    c.number_format = MONEY_FMT
    c.border = BORDER

    row += 2
    ws.cell(row=row, column=1, value='DEPARTMENT BREAKDOWN').fill = SUB_FILL
    ws.cell(row=row, column=1).font = SUB_FONT
    ws.merge_cells(f'A{row}:D{row}')
    row += 1
    set_header_row(ws, row, ['Department', 'Assessments', '', ''])
    row += 1
    for dept, label in [('physiotherapy', 'Physiotherapy'), ('nutrition', 'Nutrition'), ('general', 'General Medicine')]:
        cnt = Assessment.objects.filter(department=dept, assessment_date__date__range=[start_date, end_date]).count()
        ws.cell(row=row, column=1, value=label).border = BORDER
        c = ws.cell(row=row, column=2, value=cnt)
        c.number_format = NUM_FMT
        c.border = BORDER
        row += 1

    auto_width(ws)

    # ════════════════════════════════════════════════════════════════
    # Sheet 2 — Trend Data
    # ════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Trend Data')
    ws2.sheet_properties.tabColor = '2E8B57'

    ws2.merge_cells('A1:D1')
    ws2['A1'].value = f'Trend Data — {period_label}'
    ws2['A1'].font  = Font(bold=True, size=13, color='2E8B57')
    ws2['A1'].alignment = CENTRE

    set_header_row(ws2, 2, ['Period', 'New Patients', 'Assessments', 'Revenue (UGX)'])

    def build_trend_rows():
        rows = []
        if span_days <= 1:
            rows.append((start_date.strftime('%b %d, %Y'), _patients_in(start_date, end_date),
                         _assessments_in(start_date, end_date), _revenue_in(start_date, end_date)))
        elif span_days <= 31:
            d = start_date
            while d <= end_date:
                rows.append((d.strftime('%b %d'), _patients_in(d, d), _assessments_in(d, d), _revenue_in(d, d)))
                d += timedelta(days=1)
        elif span_days <= 90:
            d = start_date
            while d <= end_date:
                we = min(d + timedelta(days=6), end_date)
                label = f"{d.strftime('%m/%d')}–{we.strftime('%m/%d')}"
                rows.append((label, _patients_in(d, we), _assessments_in(d, we), _revenue_in(d, we)))
                d += timedelta(days=7)
        elif period == 'annual':
            for m in range(1, 13):
                _, ml = monthrange(start_date.year, m)
                ms = date(start_date.year, m, 1)
                me = date(start_date.year, m, ml)
                rows.append((ms.strftime('%b %Y'), _patients_in(ms, me), _assessments_in(ms, me), _revenue_in(ms, me)))
        else:
            cur = start_date.replace(day=1)
            while cur <= end_date:
                _, ml = monthrange(cur.year, cur.month)
                me = date(cur.year, cur.month, ml)
                be = min(me, end_date)
                rows.append((cur.strftime('%b %Y'), _patients_in(cur, be), _assessments_in(cur, be), _revenue_in(cur, be)))
                cur = date(cur.year + 1, 1, 1) if cur.month == 12 else date(cur.year, cur.month + 1, 1)
        return rows

    for r, (lbl, np_, na, rev) in enumerate(build_trend_rows(), 3):
        ws2.cell(row=r, column=1, value=lbl).border = BORDER
        c2 = ws2.cell(row=r, column=2, value=np_); c2.number_format = NUM_FMT; c2.border = BORDER
        c3 = ws2.cell(row=r, column=3, value=na);  c3.number_format = NUM_FMT; c3.border = BORDER
        c4 = ws2.cell(row=r, column=4, value=rev); c4.number_format = MONEY_FMT; c4.border = BORDER

    auto_width(ws2)

    # ════════════════════════════════════════════════════════════════
    # Sheet 3 — New vs Returning
    # ════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('New vs Returning')
    ws3.sheet_properties.tabColor = '0891B2'

    ws3.merge_cells('A1:C1')
    ws3['A1'].value = f'New vs Returning Patients — {period_label}'
    ws3['A1'].font  = Font(bold=True, size=13, color='0891B2')
    ws3['A1'].alignment = CENTRE

    set_header_row(ws3, 2, ['Period', 'First Visits (New)', 'Follow-up Visits (Returning)'])

    def build_retention_rows():
        rows = []
        if span_days <= 31:
            d = start_date
            while d <= end_date:
                nv, fup = _nv_fup(d, d)
                rows.append((d.strftime('%b %d'), nv, fup))
                d += timedelta(days=1)
        elif span_days <= 90:
            d = start_date
            while d <= end_date:
                we = min(d + timedelta(days=6), end_date)
                nv, fup = _nv_fup(d, we)
                rows.append((f"{d.strftime('%m/%d')}–{we.strftime('%m/%d')}", nv, fup))
                d += timedelta(days=7)
        elif period == 'annual':
            for m in range(1, 13):
                _, ml = monthrange(start_date.year, m)
                ms = date(start_date.year, m, 1)
                me = date(start_date.year, m, ml)
                nv, fup = _nv_fup(ms, me)
                rows.append((ms.strftime('%b %Y'), nv, fup))
        else:
            cur = start_date.replace(day=1)
            while cur <= end_date:
                _, ml = monthrange(cur.year, cur.month)
                me = min(date(cur.year, cur.month, ml), end_date)
                nv, fup = _nv_fup(cur, me)
                rows.append((cur.strftime('%b %Y'), nv, fup))
                cur = date(cur.year + 1, 1, 1) if cur.month == 12 else date(cur.year, cur.month + 1, 1)
        return rows

    total_nv = total_fup = 0
    for r, (lbl, nv, fup) in enumerate(build_retention_rows(), 3):
        ws3.cell(row=r, column=1, value=lbl).border = BORDER
        c2 = ws3.cell(row=r, column=2, value=nv);  c2.number_format = NUM_FMT; c2.border = BORDER
        c3 = ws3.cell(row=r, column=3, value=fup); c3.number_format = NUM_FMT; c3.border = BORDER
        total_nv += nv; total_fup += fup

    # Totals row
    tr = r + 1
    ws3.cell(row=tr, column=1, value='TOTAL').font = BOLD_FONT
    ws3.cell(row=tr, column=1).fill = SUB_FILL
    ws3.cell(row=tr, column=1).border = BORDER
    c2 = ws3.cell(row=tr, column=2, value=total_nv);  c2.font = BOLD_FONT; c2.number_format = NUM_FMT; c2.fill = SUB_FILL; c2.border = BORDER
    c3 = ws3.cell(row=tr, column=3, value=total_fup); c3.font = BOLD_FONT; c3.number_format = NUM_FMT; c3.fill = SUB_FILL; c3.border = BORDER

    auto_width(ws3)

    # ════════════════════════════════════════════════════════════════
    # Sheet 4 — Department Breakdown
    # ════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Department Breakdown')
    ws4.sheet_properties.tabColor = 'D97706'

    ws4.merge_cells('A1:E1')
    ws4['A1'].value = f'Department Breakdown — {period_label}'
    ws4['A1'].font  = Font(bold=True, size=13, color='D97706')
    ws4['A1'].alignment = CENTRE

    set_header_row(ws4, 2, ['Department', 'Total Assessments', 'First Visits', 'Follow-ups', 'Follow-up Required'])

    depts = [('physiotherapy', 'Physiotherapy'), ('nutrition', 'Nutrition'), ('general', 'General Medicine')]
    total_dept = [0, 0, 0, 0]
    for r, (dept, label) in enumerate(depts, 3):
        qs = Assessment.objects.filter(department=dept, assessment_date__date__range=[start_date, end_date])
        total_a = qs.count()
        fv      = qs.filter(assessment_type='first_visit').count()
        fu      = qs.filter(assessment_type='follow_up').count()
        fur     = qs.filter(follow_up_required=True).count()
        vals = [label, total_a, fv, fu, fur]
        for col, val in enumerate(vals, 1):
            c = ws4.cell(row=r, column=col, value=val)
            c.border = BORDER
            if col > 1:
                c.number_format = NUM_FMT
        total_dept[0] += total_a; total_dept[1] += fv; total_dept[2] += fu; total_dept[3] += fur

    tr = r + 1
    for col, val in enumerate(['TOTAL'] + total_dept, 1):
        c = ws4.cell(row=tr, column=col, value=val)
        c.font = BOLD_FONT; c.fill = SUB_FILL; c.border = BORDER
        if col > 1:
            c.number_format = NUM_FMT

    auto_width(ws4)

    # ════════════════════════════════════════════════════════════════
    # Sheet 5 — Most Frequent Patients
    # ════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet('Top Patients')
    ws5.sheet_properties.tabColor = '7C3AED'

    ws5.merge_cells('A1:F1')
    ws5['A1'].value = f'Most Frequent Patients — {period_label}'
    ws5['A1'].font  = Font(bold=True, size=13, color='7C3AED')
    ws5['A1'].alignment = CENTRE

    set_header_row(ws5, 2, ['#', 'Patient ID', 'Full Name', 'Phone', 'Visits in Period', 'Last Visit Date'])

    top_patients = Patient.objects.filter(
        is_active=True,
        assessments__assessment_date__date__range=[start_date, end_date]
    ).annotate(
        visit_count=Count('assessments', filter=Q(
            assessments__assessment_date__date__range=[start_date, end_date]
        )),
        last_visit=Max('assessments__assessment_date')
    ).filter(visit_count__gt=0).order_by('-visit_count')[:50]

    for r, p in enumerate(top_patients, 3):
        ws5.cell(row=r, column=1, value=r - 2).border = BORDER
        ws5.cell(row=r, column=2, value=p.patient_id).border = BORDER
        ws5.cell(row=r, column=3, value=p.get_full_name()).border = BORDER
        ws5.cell(row=r, column=4, value=getattr(p, 'phone', '')).border = BORDER
        c5 = ws5.cell(row=r, column=5, value=p.visit_count); c5.number_format = NUM_FMT; c5.border = BORDER
        lv = p.last_visit.date() if p.last_visit else ''
        ws5.cell(row=r, column=6, value=str(lv) if lv else '').border = BORDER

    auto_width(ws5)

    # ════════════════════════════════════════════════════════════════
    # Sheet 6 — Appointments Detail
    # ════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet('Appointments')
    ws6.sheet_properties.tabColor = '059669'

    ws6.merge_cells('A1:E1')
    ws6['A1'].value = f'Appointments — {period_label}'
    ws6['A1'].font  = Font(bold=True, size=13, color='059669')
    ws6['A1'].alignment = CENTRE

    set_header_row(ws6, 2, ['Date', 'Patient', 'Service', 'Provider', 'Status'])

    appts = Appointment.objects.filter(
        appointment_date__range=[start_date, end_date]
    ).select_related('patient', 'service', 'provider').order_by('appointment_date', 'appointment_time')

    status_fills = {
        'completed':  PatternFill('solid', fgColor='D1FAE5'),
        'cancelled':  PatternFill('solid', fgColor='FEE2E2'),
        'no_show':    PatternFill('solid', fgColor='FEF3C7'),
        'scheduled':  PatternFill('solid', fgColor='EFF6FF'),
    }

    for r, appt in enumerate(appts, 3):
        appt_date = appt.appointment_date.strftime('%b %d, %Y') if appt.appointment_date else ''
        patient_name = appt.patient.get_full_name() if appt.patient else ''
        service_name = appt.service.name if appt.service else ''
        provider_name = f"{appt.provider.get_full_name()}" if appt.provider else ''
        status = appt.get_status_display() if hasattr(appt, 'get_status_display') else appt.status

        row_fill = status_fills.get(appt.status, PatternFill())
        for col, val in enumerate([appt_date, patient_name, service_name, provider_name, status], 1):
            c = ws6.cell(row=r, column=col, value=val)
            c.border = BORDER
            c.fill   = row_fill

    auto_width(ws6)

    # ── Build HTTP response ──────────────────────────────────────────
    filename = f"statistics_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def clinical_summary_report(request):
    """Comprehensive clinical summary report across all departments"""
    from patients.models import Assessment, VitalSigns
    
    # Get filter parameters
    start_date = request.GET.get('start_date', date.today().replace(day=1).isoformat())
    end_date = request.GET.get('end_date', date.today().isoformat())
    
    start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Assessment statistics by department
    physio_count = Assessment.objects.filter(
        department='physiotherapy',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    nutrition_count = Assessment.objects.filter(
        department='nutrition',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    general_count = Assessment.objects.filter(
        department='general',
        assessment_date__range=[start_date, end_date]
    ).count()
    
    # Vital signs monitoring
    vital_signs_count = VitalSigns.objects.filter(
        recorded_date__range=[start_date, end_date]
    ).count()
    
    # Patient engagement
    unique_patients = Assessment.objects.filter(
        assessment_date__range=[start_date, end_date]
    ).values('patient').distinct().count()
    
    # Department distribution
    dept_labels = ['Physiotherapy', 'Nutrition', 'General']
    dept_data = [physio_count, nutrition_count, general_count]
    
    # Assessment type distribution
    first_visits = Assessment.objects.filter(
        assessment_date__range=[start_date, end_date],
        assessment_type='first_visit'
    ).count()
    
    follow_ups = Assessment.objects.filter(
        assessment_date__range=[start_date, end_date],
        assessment_type='follow_up'
    ).count()
    
    # Recent assessments across all departments
    recent_assessments = Assessment.objects.filter(
        assessment_date__range=[start_date, end_date]
    ).select_related('patient', 'assessed_by').order_by('-assessment_date')[:30]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'physio_count': physio_count,
        'nutrition_count': nutrition_count,
        'general_count': general_count,
        'vital_signs_count': vital_signs_count,
        'unique_patients': unique_patients,
        'first_visits': first_visits,
        'follow_ups': follow_ups,
        'dept_labels': json.dumps(dept_labels),
        'dept_data': json.dumps(dept_data),
        'recent_assessments': recent_assessments,
    }
    return render(request, 'reports/clinical_summary.html', context)
