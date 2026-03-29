import time
import os
from datetime import datetime, timedelta
from django.utils import timezone
from django.http import HttpResponse
from django.conf import settings
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
import csv
import io
from .models import ReportAuditLog, ReportExport

class ReportAuditMixin:
    """Mixin to add audit logging to report views"""
    
    def log_report_activity(self, request, report_type, report_name, action, 
                          execution_time=0, record_count=0, file_size=0, 
                          success=True, error_message='', parameters=None):
        """Log report activity for audit purposes"""
        try:
            ReportAuditLog.objects.create(
                user=request.user,
                report_type=report_type,
                report_name=report_name,
                action=action,
                parameters=parameters or {},
                execution_time=execution_time,
                record_count=record_count,
                file_size=file_size,
                ip_address=self.get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                success=success,
                error_message=error_message
            )
        except Exception as e:
            # Don't let audit logging break the main functionality
            print(f"Audit logging failed: {e}")
    
    def get_client_ip(self, request):
        """Get client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip

class PDFReportGenerator:
    """Professional PDF report generator"""
    
    def __init__(self, title, subtitle="", author="Physio Nutrition Clinic"):
        self.title = title
        self.subtitle = subtitle
        self.author = author
        self.styles = getSampleStyleSheet()
        self.buffer = io.BytesIO()
        
        # Custom styles
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#2c3e50'),
            alignment=1  # Center
        )
        
        self.subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            textColor=colors.HexColor('#34495e'),
            alignment=1  # Center
        )
        
        self.header_style = ParagraphStyle(
            'CustomHeader',
            parent=self.styles['Heading3'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#2980b9'),
            spaceBefore=20
        )
    
    def create_document(self):
        """Create the PDF document"""
        self.doc = SimpleDocTemplate(
            self.buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        return self.doc
    
    def add_header(self, story):
        """Add header to the document"""
        # Title
        story.append(Paragraph(self.title, self.title_style))
        
        # Subtitle
        if self.subtitle:
            story.append(Paragraph(self.subtitle, self.subtitle_style))
        
        # Date and time
        date_str = datetime.now().strftime("%B %d, %Y at %I:%M %p")
        date_para = Paragraph(f"Generated on {date_str}", self.styles['Normal'])
        story.append(date_para)
        story.append(Spacer(1, 20))
    
    def add_section(self, story, title, content):
        """Add a section to the document"""
        story.append(Paragraph(title, self.header_style))
        
        if isinstance(content, str):
            story.append(Paragraph(content, self.styles['Normal']))
        elif isinstance(content, list):
            for item in content:
                if isinstance(item, dict):
                    # Handle table data
                    self.add_table(story, item)
                else:
                    story.append(Paragraph(str(item), self.styles['Normal']))
        
        story.append(Spacer(1, 12))
    
    def add_table(self, story, table_data):
        """Add a table to the document"""
        if 'headers' in table_data and 'data' in table_data:
            # Prepare table data
            data = [table_data['headers']] + table_data['data']
            
            # Create table
            table = Table(data, repeatRows=1)
            
            # Style the table
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table)
    
    def add_summary_stats(self, story, stats):
        """Add summary statistics section"""
        story.append(Paragraph("Summary Statistics", self.header_style))
        
        # Create a 2-column table for stats
        stats_data = []
        for key, value in stats.items():
            stats_data.append([key.replace('_', ' ').title(), str(value)])
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))
    
    def generate(self, content_data):
        """Generate the PDF report"""
        story = []
        
        # Add header
        self.add_header(story)
        
        # Add summary statistics if provided
        if 'summary_stats' in content_data:
            self.add_summary_stats(story, content_data['summary_stats'])
        
        # Add sections
        if 'sections' in content_data:
            for section in content_data['sections']:
                self.add_section(story, section['title'], section['content'])
        
        # Build PDF
        doc = self.create_document()
        doc.build(story)
        
        # Get PDF data
        pdf_data = self.buffer.getvalue()
        self.buffer.close()
        
        return pdf_data

class ExcelReportGenerator:
    """Professional Excel report generator"""
    
    def __init__(self, title, subtitle=""):
        self.title = title
        self.subtitle = subtitle
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.current_row = 1
        
        # Define styles
        self.title_font = Font(name='Arial', size=16, bold=True, color='2C3E50')
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.data_font = Font(name='Arial', size=10)
        self.header_fill = PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid')
        self.alt_fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
        
        # Border style
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def add_title(self):
        """Add title to the worksheet"""
        self.worksheet.cell(row=self.current_row, column=1, value=self.title)
        self.worksheet.cell(row=self.current_row, column=1).font = self.title_font
        self.current_row += 1
        
        if self.subtitle:
            self.worksheet.cell(row=self.current_row, column=1, value=self.subtitle)
            self.worksheet.cell(row=self.current_row, column=1).font = Font(name='Arial', size=12, color='34495E')
            self.current_row += 1
        
        # Add generation date
        date_str = datetime.now().strftime("%B %d, %Y at %I:%M %p")
        self.worksheet.cell(row=self.current_row, column=1, value=f"Generated on {date_str}")
        self.worksheet.cell(row=self.current_row, column=1).font = Font(name='Arial', size=10, italic=True)
        self.current_row += 2
    
    def add_summary_section(self, stats):
        """Add summary statistics section"""
        self.worksheet.cell(row=self.current_row, column=1, value="Summary Statistics")
        self.worksheet.cell(row=self.current_row, column=1).font = Font(name='Arial', size=14, bold=True)
        self.current_row += 1
        
        for key, value in stats.items():
            self.worksheet.cell(row=self.current_row, column=1, value=key.replace('_', ' ').title())
            self.worksheet.cell(row=self.current_row, column=2, value=value)
            self.worksheet.cell(row=self.current_row, column=1).font = Font(name='Arial', size=10, bold=True)
            self.worksheet.cell(row=self.current_row, column=2).font = self.data_font
            self.current_row += 1
        
        self.current_row += 1
    
    def add_table(self, title, headers, data):
        """Add a data table to the worksheet"""
        # Add table title
        self.worksheet.cell(row=self.current_row, column=1, value=title)
        self.worksheet.cell(row=self.current_row, column=1).font = Font(name='Arial', size=12, bold=True)
        self.current_row += 1
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=self.current_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        self.current_row += 1
        
        # Add data rows
        for row_data in data:
            for col, value in enumerate(row_data, 1):
                cell = self.worksheet.cell(row=self.current_row, column=col, value=value)
                cell.font = self.data_font
                cell.border = self.thin_border
                
                # Alternate row coloring
                if (self.current_row - len(headers) - 1) % 2 == 0:
                    cell.fill = self.alt_fill
            
            self.current_row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            self.worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].auto_size = True
        
        self.current_row += 1
    
    def add_chart(self, chart_type, title, data_range, categories_range=None):
        """Add a chart to the worksheet"""
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'line':
            chart = LineChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            return
        
        chart.title = title
        chart.add_data(data_range, titles_from_data=True)
        
        if categories_range:
            chart.set_categories(categories_range)
        
        # Position chart
        chart_cell = f"A{self.current_row}"
        self.worksheet.add_chart(chart, chart_cell)
        self.current_row += 15  # Leave space for chart
    
    def generate(self, content_data):
        """Generate the Excel report"""
        # Add title
        self.add_title()
        
        # Add summary statistics if provided
        if 'summary_stats' in content_data:
            self.add_summary_section(content_data['summary_stats'])
        
        # Add tables
        if 'tables' in content_data:
            for table in content_data['tables']:
                self.add_table(table['title'], table['headers'], table['data'])
        
        # Save to buffer
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()

def create_report_export(user, report_type, report_name, export_format, content_data, parameters=None):
    """Create and save a report export"""
    try:
        # Generate the export based on format
        if export_format == 'pdf':
            generator = PDFReportGenerator(report_name)
            file_data = generator.generate(content_data)
            content_type = 'application/pdf'
            file_extension = 'pdf'
        elif export_format == 'excel':
            generator = ExcelReportGenerator(report_name)
            file_data = generator.generate(content_data)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            file_extension = 'xlsx'
        elif export_format == 'csv':
            # Simple CSV export for tabular data
            buffer = io.StringIO()
            if 'tables' in content_data and content_data['tables']:
                table = content_data['tables'][0]  # Use first table
                writer = csv.writer(buffer)
                writer.writerow(table['headers'])
                writer.writerows(table['data'])
            file_data = buffer.getvalue().encode('utf-8')
            content_type = 'text/csv'
            file_extension = 'csv'
        else:
            raise ValueError(f"Unsupported export format: {export_format}")
        
        # Create export record
        export = ReportExport.objects.create(
            user=user,
            report_type=report_type,
            report_name=report_name,
            export_format=export_format,
            file_size=len(file_data),
            parameters=parameters or {},
            expires_at=timezone.now() + timedelta(days=7)  # Expire after 7 days
        )
        
        # Create HTTP response
        filename = f"{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_extension}"
        response = HttpResponse(file_data, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response, export
        
    except Exception as e:
        raise Exception(f"Failed to create export: {str(e)}")

def send_scheduled_report(scheduled_report, content_data):
    """Send a scheduled report via email"""
    try:
        # Generate PDF for email attachment
        generator = PDFReportGenerator(
            scheduled_report.configuration.name,
            f"Scheduled Report - {scheduled_report.configuration.get_report_type_display()}"
        )
        pdf_data = generator.generate(content_data)
        
        # Create email
        subject = f"Scheduled Report: {scheduled_report.configuration.name}"
        message = render_to_string('reports/email/scheduled_report.html', {
            'report_name': scheduled_report.configuration.name,
            'report_type': scheduled_report.configuration.get_report_type_display(),
            'generated_at': timezone.now(),
        })
        
        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=scheduled_report.recipients
        )
        
        # Attach PDF
        filename = f"{scheduled_report.configuration.name}_{datetime.now().strftime('%Y%m%d')}.pdf"
        email.attach(filename, pdf_data, 'application/pdf')
        
        # Send email
        email.send()
        
        return True
        
    except Exception as e:
        print(f"Failed to send scheduled report: {e}")
        return False

def get_report_performance_metrics():
    """Get performance metrics for reports"""
    from django.db.models import Avg, Count, Sum
    
    # Get audit log statistics
    total_reports = ReportAuditLog.objects.filter(action='generated').count()
    avg_execution_time = ReportAuditLog.objects.filter(action='generated').aggregate(
        avg_time=Avg('execution_time')
    )['avg_time'] or 0
    
    # Most popular reports
    popular_reports = ReportAuditLog.objects.filter(action='generated').values(
        'report_type', 'report_name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Export statistics
    export_stats = ReportExport.objects.values('export_format').annotate(
        count=Count('id'),
        total_size=Sum('file_size')
    )
    
    return {
        'total_reports_generated': total_reports,
        'average_execution_time': round(avg_execution_time, 2),
        'popular_reports': list(popular_reports),
        'export_statistics': list(export_stats),
    }
